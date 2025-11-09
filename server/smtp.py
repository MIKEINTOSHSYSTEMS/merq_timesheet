# server/smtp.py
# SMTP Configuration and Email Service for MERQ Timesheet System
# Version: 1.0.0
# Author: Michael Kifle Teferra
# Date: November 2025

# Required Libraries
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import pandas as pd
from datetime import datetime, timedelta
import sys
import tempfile
import subprocess
from packaging import version
import os
from PIL import Image, ImageTk, ImageDraw, ImageFont
import math
import requests
import json
from typing import Tuple, Optional, Dict, List
import threading
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import webbrowser
import sqlite3
import hashlib
# import bcrypt  # Commented out to avoid import issues
import re
import smtplib
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from functools import wraps
from io import BytesIO

# Add the src directory to Python path to import timesheet.py
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

# Import EthiopianDateConverter directly from the backup file to avoid dependency issues
import importlib.util
spec = importlib.util.spec_from_file_location("ethiopian_converter", os.path.join(os.path.dirname(__file__), '..', 'src', 'bak.timesheet.py'))
ethiopian_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(ethiopian_module)
EthiopianDateConverter = ethiopian_module.EthiopianDateConverter

class SMTPConfig:
    """SMTP Configuration Manager"""
    
    def __init__(self):
        self.config = {
            'SMTPServer': 'cloud.merqconsultancy.org',
            'SMTPPort': 587,
            'SMTPUser': 'lifebox@cloud.merqconsultancy.org',
            'SMTPPassword': 'LifeboxCloud',
            'UseTLS': False
            #'UseTLS': True
        }
        
        # Setup logging
        self.setup_logging()
    
    def setup_logging(self):
        """Setup logging configuration"""
        log_dir = 'server'
        os.makedirs(log_dir, exist_ok=True)
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(os.path.join(log_dir, 'smtp.log')),
                logging.StreamHandler()
            ]
        )
        
        self.logger = logging.getLogger('MERQ_SMTP')
    
    def get_config(self):
        """Get SMTP configuration"""
        return self.config
    
    def update_config(self, new_config):
        """Update SMTP configuration"""
        self.config.update(new_config)
        self.logger.info("SMTP configuration updated")

class EmailService:
    """Email Service for sending timesheet emails"""
    
    def __init__(self):
        self.smtp_config = SMTPConfig()
        self.config = self.smtp_config.get_config()
        self.logger = self.smtp_config.logger
    
    def send_timesheet_email(self, timesheet_file, user_session, hr_users):
        """
        Send timesheet via email to HR with CC to user
        
        Args:
            timesheet_file: Path to the Excel timesheet file
            user_session: UserSession object with user data
            hr_users: List of HR users (position_id = 18)
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get current Ethiopian date
            eth_converter = EthiopianDateConverter()
            current_eth_date = eth_converter.get_current_ethiopian_date()
            
            month_name = current_eth_date['month_name']
            year = current_eth_date['year']
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.config['SMTPUser']
            
            # Set recipients - HR users and CC the sender
            hr_emails = [hr['email'] for hr in hr_users if hr.get('email')]
            if not hr_emails:
                hr_emails = ['haymanot.a@merqconsultancy.org']  # Fallback
            
            msg['To'] = ', '.join(hr_emails)
            msg['Cc'] = user_session.email
            msg['Subject'] = f"MERQ Timesheet for {month_name} {year} - {user_session.full_name}"
            
            # Email body
            body = self._create_email_body(user_session, month_name, year, hr_users)
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach timesheet file
            if os.path.exists(timesheet_file):
                with open(timesheet_file, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                
                encoders.encode_base64(part)
                filename = os.path.basename(timesheet_file)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {filename}'
                )
                msg.attach(part)
            
            # Send email
            all_recipients = hr_emails + [user_session.email]
            success = self._send_email(msg, all_recipients)
            
            if success:
                self.logger.info(f"Timesheet email sent successfully for {user_session.full_name}")
                self.logger.info(f"Recipients: HR={hr_emails}, CC={user_session.email}")
            else:
                self.logger.error(f"Failed to send timesheet email for {user_session.full_name}")
            
            return success
            
        except Exception as e:
            self.logger.error(f"Error sending timesheet email: {str(e)}")
            return False
    
    def _create_email_body(self, user_session, month_name, year, hr_users):
        """Create email body content"""
        
        hr_names = [hr.get('full_name', 'HR Department') for hr in hr_users]
        hr_recipients = ", ".join(hr_names)
        
        body = f"""
Dear {hr_recipients},

Please find attached the timesheet for your review and approval.

Employee Details:
- Name: {user_session.full_name}
- Position: {user_session.position}
- Department: {user_session.department}
- Employee ID: {user_session.employee_id}
- Supervisor: {user_session.supervisor_name}
- Supervisor Position: {user_session.supervisor_position_title}

Timesheet Period: {month_name} {year}

This timesheet has been generated and submitted through the MERQ Timesheet System.

Summary:
- Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- Sent to: HR Department
- CC: {user_session.email}

Best regards,
{user_session.full_name}
{user_session.position}
MERQ Consultancy PLC

---
This is an automated email from MERQ Timesheet System.
"""
        return body
    
    def _send_email(self, msg, recipients):
        """Send the actual email"""
        try:
            server = smtplib.SMTP(self.config['SMTPServer'], self.config['SMTPPort'])
            
            if self.config['UseTLS']:
                server.starttls()
            
            server.login(self.config['SMTPUser'], self.config['SMTPPassword'])
            text = msg.as_string()
            server.sendmail(self.config['SMTPUser'], recipients, text)
            server.quit()
            
            return True
            
        except Exception as e:
            self.logger.error(f"SMTP error: {str(e)}")
            return False
    
    def test_connection(self):
        """Test SMTP connection"""
        try:
            server = smtplib.SMTP(self.config['SMTPServer'], self.config['SMTPPort'])
            
            if self.config['UseTLS']:
                server.starttls()
            
            server.login(self.config['SMTPUser'], self.config['SMTPPassword'])
            server.quit()
            
            self.logger.info("SMTP connection test successful")
            return True
            
        except Exception as e:
            self.logger.error(f"SMTP connection test failed: {str(e)}")
            return False

# Global instance
email_service = EmailService()