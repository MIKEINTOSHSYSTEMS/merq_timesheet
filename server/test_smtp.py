# test_smtp.py
# SMTP Configuration Test Tool with GUI
# Version: 1.0.0
# Author: Michael Kifle Teferra
# Date: November 2025

# Required Libraries
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import pandas as pd
from datetime import datetime, timedelta
import sys
import tempfile
import subprocess
from packaging import version
import os
from PIL import Image, ImageTk, ImageDraw, ImageFont
import math
import requests
import json
from typing import Tuple, Optional, Dict, List
import threading
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import webbrowser
import sqlite3
import hashlib
import bcrypt
import re
import smtplib
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from functools import wraps
from io import BytesIO

# Add the server directory to path
server_dir = os.path.join(os.path.dirname(__file__), 'server')
if server_dir not in sys.path:
    sys.path.append(server_dir)

from smtp import EmailService, SMTPConfig

class SMTPTester:
    def __init__(self, root):
        self.root = root
        self.root.title("MERQ SMTP Configuration Tester")
        self.root.geometry("800x600")
        
        self.email_service = EmailService()
        self.create_widgets()
    
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="SMTP Configuration Tester", 
                               font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Configuration frame
        config_frame = ttk.LabelFrame(main_frame, text="SMTP Configuration", padding="10")
        config_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Server configuration
        ttk.Label(config_frame, text="SMTP Server:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.server_var = tk.StringVar(value="cloud.merqconsultancy.org")
        ttk.Entry(config_frame, textvariable=self.server_var, width=30).grid(row=0, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(config_frame, text="Port:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.port_var = tk.StringVar(value="587")
        ttk.Entry(config_frame, textvariable=self.port_var, width=10).grid(row=1, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(config_frame, text="Username:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.user_var = tk.StringVar(value="app@cloud.merqconsultancy.org")
        ttk.Entry(config_frame, textvariable=self.user_var, width=30).grid(row=2, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(config_frame, text="Password:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.password_var = tk.StringVar(value="MerqAppCloud")
        ttk.Entry(config_frame, textvariable=self.password_var, show="*", width=30).grid(row=3, column=1, sticky=tk.W, pady=2)
        
        self.use_tls_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, text="Use TLS", variable=self.use_tls_var).grid(row=4, column=1, sticky=tk.W, pady=2)
        
        # Test email frame
        email_frame = ttk.LabelFrame(main_frame, text="Test Email", padding="10")
        email_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(email_frame, text="To Email:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.to_email_var = tk.StringVar(value="support@merqconsultancy.org")
        ttk.Entry(email_frame, textvariable=self.to_email_var, width=30).grid(row=0, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(email_frame, text="Subject:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.subject_var = tk.StringVar(value="SMTP Test Email")
        ttk.Entry(email_frame, textvariable=self.subject_var, width=30).grid(row=1, column=1, sticky=tk.W, pady=2)
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(button_frame, text="Test Connection", 
                  command=self.test_connection).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Send Test Email", 
                  command=self.send_test_email).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Load Defaults", 
                  command=self.load_defaults).pack(side=tk.LEFT)
        
        # Log frame
        log_frame = ttk.LabelFrame(main_frame, text="Log Output", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=80)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Clear log button
        ttk.Button(main_frame, text="Clear Log", 
                  command=self.clear_log).pack(pady=(10, 0))
    
    def log_message(self, message):
        """Add message to log"""
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def clear_log(self):
        """Clear log messages"""
        self.log_text.delete(1.0, tk.END)
    
    def load_defaults(self):
        """Load default configuration"""
        self.server_var.set("cloud.merqconsultancy.org")
        self.port_var.set("587")
        self.user_var.set("app@cloud.merqconsultancy.org")
        self.password_var.set("MerqAppCloud")
        self.use_tls_var.set(True)
        self.log_message("Loaded default configuration")
    
    def test_connection(self):
        """Test SMTP connection in background thread"""
        def test_thread():
            self.log_message("Testing SMTP connection...")
            
            # Update configuration
            config = {
                'SMTPServer': self.server_var.get(),
                'SMTPPort': int(self.port_var.get()),
                'SMTPUser': self.user_var.get(),
                'SMTPPassword': self.password_var.get(),
                'UseTLS': self.use_tls_var.get()
            }
            
            self.email_service.smtp_config.update_config(config)
            
            success = self.email_service.test_connection()
            
            if success:
                self.log_message("✅ SMTP connection test SUCCESSFUL")
                messagebox.showinfo("Success", "SMTP connection test successful!")
            else:
                self.log_message("❌ SMTP connection test FAILED")
                messagebox.showerror("Error", "SMTP connection test failed!")
        
        thread = threading.Thread(target=test_thread, daemon=True)
        thread.start()
    
    def send_test_email(self):
        """Send test email in background thread"""
        def send_thread():
            self.log_message("Sending test email...")
            
            # Update configuration
            config = {
                'SMTPServer': self.server_var.get(),
                'SMTPPort': int(self.port_var.get()),
                'SMTPUser': self.user_var.get(),
                'SMTPPassword': self.password_var.get(),
                'UseTLS': self.use_tls_var.get()
            }
            
            self.email_service.smtp_config.update_config(config)
            
            # Create test email
            try:
                import smtplib
                from email.mime.text import MIMEText
                from email.mime.multipart import MIMEMultipart

                msg = MIMEMultipart()
                msg['From'] = self.user_var.get()
                msg['To'] = self.to_email_var.get()
                msg['Subject'] = self.subject_var.get()

                sent_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                body = f"""
                This is a test email from MERQ SMTP Tester.

                Configuration:
                - Server: {self.server_var.get()}
                - Port: {self.port_var.get()}
                - User: {self.user_var.get()}
                - TLS: {self.use_tls_var.get()}

                Sent at: {sent_at}
                """
                
                msg.attach(MIMEText(body, 'plain'))
                
                server = smtplib.SMTP(self.server_var.get(), int(self.port_var.get()))
                
                if self.use_tls_var.get():
                    server.starttls()
                
                server.login(self.user_var.get(), self.password_var.get())
                text = msg.as_string()
                server.sendmail(self.user_var.get(), [self.to_email_var.get()], text)
                server.quit()
                
                self.log_message("✅ Test email sent SUCCESSFULLY")
                messagebox.showinfo("Success", "Test email sent successfully!")
                
            except Exception as e:
                self.log_message(f"❌ Failed to send test email: {str(e)}")
                messagebox.showerror("Error", f"Failed to send test email: {str(e)}")
        
        thread = threading.Thread(target=send_thread, daemon=True)
        thread.start()

def main():
    root = tk.Tk()
    app = SMTPTester(root)
    root.mainloop()

if __name__ == "__main__":
    main()