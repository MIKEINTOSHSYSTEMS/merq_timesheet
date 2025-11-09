# webapp.py - MERQ Timesheet Web Application
# This web application leverages the existing timesheet.py functionality
# while providing a modern, responsive web interface.

# MERQ Timesheet Management System
# Version: 1.0.0.1
# Author: Michael Kifle Teferra
# Date: Novmeber 2025
# Description: Core functionalities for MERQ Timesheet application including
#              database management, user authentication, email handling, excel formatting,
#              and Ethiopian date conversion.
# License: MIT License

# Required Libraries
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import pandas as pd
from datetime import datetime, timedelta
import sys
import tempfile
import subprocess
from packaging import version
import os
from PIL import Image, ImageTk, ImageDraw, ImageFont
import math
import requests
import json
from typing import Tuple, Optional, Dict, List
import threading
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import webbrowser
import sqlite3
import hashlib
import bcrypt
import re
import smtplib
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from functools import wraps
from io import BytesIO

# Add the src directory to Python path to import timesheet.py
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

# Import Flask and related modules
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, send_file
from flask_session import Session
import pandas as pd

# Import existing classes from timesheet.py
try:
    from timesheet import (
        DatabaseManager, UserSession, EthiopianDateConverter, 
        ModernTheme, ExcelFormatter, EmailManager, TimesheetApp
    )
except ImportError as e:
    print(f"Error importing from timesheet.py: {e}")
    print("Make sure timesheet.py is in the src directory")
    sys.exit(1)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Flask application
app = Flask(__name__)
app.secret_key = 'merq_timesheet_web_app_secret_key_2025'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
Session(app)

# Global instances
db_manager = DatabaseManager()
ethiopian_converter = EthiopianDateConverter()

# In-memory storage for timesheet data (replace with database in production)
timesheet_storage = {}
user_projects = {}
user_timesheet_instances = {}

def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    """Get current user from session"""
    if 'user_id' not in session:
        return None
    
    user_data = session.get('user_data')
    if user_data:
        return UserSession(user_data)
    return None

def get_user_timesheet_key(user_id, year, month):
    """Generate a unique key for user's timesheet data"""
    return f"{user_id}_{year}_{month}"

def get_user_timesheet_instance(user_session, year, month):
    """Get or create a timesheet instance for the user"""
    timesheet_key = get_user_timesheet_key(user_session.user_id, year, month)
    
    if timesheet_key not in user_timesheet_instances:
        # Create a mock root for the timesheet app
        class MockRoot:
            def __init__(self):
                self.title = "MERQ Timesheet Web"
        
        mock_root = MockRoot()
        
        # Create timesheet instance
        timesheet_instance = type('WebTimesheetApp', (object,), {})()
        
        # Initialize with user data
        timesheet_instance.user_session = user_session
        timesheet_instance.selected_year = type('Var', (object,), {'get': lambda: year})()
        timesheet_instance.selected_month_name = ethiopian_converter.MONTHS_AMHARIC[month-1]
        timesheet_instance.projects = []
        timesheet_instance.leave_data = {
            "vacation": {'entries': {}, 'total_var': type('Var', (object,), {'get': lambda: 0.0})()},
            "sick_leave": {'entries': {}, 'total_var': type('Var', (object,), {'get': lambda: 0.0})()},
            "holiday": {'entries': {}, 'total_var': type('Var', (object,), {'get': lambda: 0.0})()},
            "personal_leave": {'entries': {}, 'total_var': type('Var', (object,), {'get': lambda: 0.0})()},
            "bereavement": {'entries': {}, 'total_var': type('Var', (object,), {'get': lambda: 0.0})()},
            "other": {'entries': {}, 'total_var': type('Var', (object,), {'get': lambda: 0.0})()}
        }
        
        # Add methods from timesheet.py
        def safe_float_convert(value):
            try:
                if value == "" or value is None:
                    return 0.0
                return float(value)
            except (ValueError, TypeError):
                return 0.0
        
        def calculate_total_working_hours():
            month_days = ethiopian_converter.get_ethiopian_month_days(year, month)
            total_hours = 0.0
            
            for day in range(1, month_days + 1):
                weekday_index = ethiopian_converter.get_ethiopian_weekday(year, month, day)
                
                if weekday_index < 4:  # Monday-Thursday: 8 hours
                    day_hours = 8.0
                elif weekday_index == 4:  # Friday: 8 hours
                    day_hours = 8.0
                elif weekday_index == 5:  # Saturday: 4 hours
                    day_hours = 4.0
                else:  # Sunday: 0 hours
                    day_hours = 0.0
                
                total_hours += day_hours
            
            return total_hours
        
        timesheet_instance.safe_float_convert = safe_float_convert
        timesheet_instance.calculate_total_working_hours = calculate_total_working_hours
        
        user_timesheet_instances[timesheet_key] = timesheet_instance
    
    return user_timesheet_instances[timesheet_key]

def initialize_user_timesheet(user_id, year, month, month_days):
    """Initialize timesheet data structure for a user"""
    timesheet_key = get_user_timesheet_key(user_id, year, month)
    
    if timesheet_key not in timesheet_storage:
        timesheet_storage[timesheet_key] = {
            'projects': {},
            'leave_entries': {
                'vacation': {day: 0.0 for day in range(1, month_days + 1)},
                'sick_leave': {day: 0.0 for day in range(1, month_days + 1)},
                'holiday': {day: 0.0 for day in range(1, month_days + 1)},
                'personal_leave': {day: 0.0 for day in range(1, month_days + 1)},
                'bereavement': {day: 0.0 for day in range(1, month_days + 1)},
                'other': {day: 0.0 for day in range(1, month_days + 1)}
            },
            'daily_totals': {day: 0.0 for day in range(1, month_days + 1)},
            'leave_totals': {day: 0.0 for day in range(1, month_days + 1)},
            'grand_totals': {day: 0.0 for day in range(1, month_days + 1)}
        }
    
    return timesheet_storage[timesheet_key]

def initialize_user_projects(user_id, year, month):
    """Initialize user projects with default MERQ Internal project"""
    if user_id not in user_projects:
        user_projects[user_id] = {}
    
    timesheet_key = get_user_timesheet_key(user_id, year, month)
    if timesheet_key not in user_projects[user_id]:
        # Calculate dynamic hours for MERQ Internal
        month_days = ethiopian_converter.get_ethiopian_month_days(year, month)
        total_hours = 0.0
        
        for day in range(1, month_days + 1):
            weekday_index = ethiopian_converter.get_ethiopian_weekday(year, month, day)
            
            if weekday_index < 4:  # Monday-Thursday: 8 hours
                day_hours = 8.0
            elif weekday_index == 4:  # Friday: 8 hours
                day_hours = 8.0
            elif weekday_index == 5:  # Saturday: 4 hours
                day_hours = 4.0
            else:  # Sunday: 0 hours
                day_hours = 0.0
            
            total_hours += day_hours
        
        user_projects[user_id][timesheet_key] = [
            {
                'id': 1,
                'name': 'MERQ Internal',
                'allocated_hours': total_hours,
                'total_hours': 0.0,
                'hours': {}
            }
        ]
    
    return user_projects[user_id][timesheet_key]

@app.route('/')
def index():
    """Home page - redirect to login or dashboard"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        
        if not email or not password:
            flash('Please enter both email and password', 'error')
            return render_template('login.html')
        
        # Normalize email
        normalized_email = normalize_email(email)
        
        # Validate credentials using existing DatabaseManager
        user = db_manager.validate_user_credentials(normalized_email, password)
        
        if user:
            # Create user session
            user_session = UserSession(user)
            
            # Store in Flask session
            session['user_id'] = user_session.user_id
            session['user_data'] = user
            session['logged_in'] = True
            session.permanent = True
            
            flash(f'Welcome back, {user_session.full_name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """User logout"""
    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard"""
    user_session = get_current_user()
    current_eth_date = ethiopian_converter.get_current_ethiopian_date()
    
    return render_template('dashboard.html',
                         user=user_session,
                         current_eth_date=current_eth_date)

@app.route('/timesheet')
@login_required
def timesheet():
    """Timesheet main page"""
    user_session = get_current_user()
    current_eth_date = ethiopian_converter.get_current_ethiopian_date()
    
    # Get available years and months
    years = list(range(2010, current_eth_date['year'] + 1))
    months = [(i+1, ethiopian_converter.MONTHS_AMHARIC[i]) for i in range(13)]
    
    return render_template('timesheet.html',
                         user=user_session,
                         current_eth_date=current_eth_date,
                         years=years,
                         months=months,
                         eth_months_amharic=ethiopian_converter.MONTHS_AMHARIC)

@app.route('/api/timesheet/data', methods=['POST'])
@login_required
def get_timesheet_data():
    """Get timesheet data for specific month/year"""
    user_session = get_current_user()
    
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    
    if not year or not month:
        return jsonify({'error': 'Year and month are required'}), 400
    
    # Calculate month details
    month_days = ethiopian_converter.get_ethiopian_month_days(year, month)
    
    # Initialize user timesheet data
    timesheet_data = initialize_user_timesheet(user_session.user_id, year, month, month_days)
    
    # Initialize user projects
    projects = initialize_user_projects(user_session.user_id, year, month)
    
    # Generate calendar data
    calendar_data = []
    for day in range(1, month_days + 1):
        weekday_index = ethiopian_converter.get_ethiopian_weekday(year, month, day)
        eth_date = ethiopian_converter.format_ethiopian_date(year, month, day)
        
        calendar_data.append({
            'day': day,
            'date': eth_date,
            'weekday_amharic': ethiopian_converter.WEEK_DAYS_AMHARIC[weekday_index],
            'weekday_english': ethiopian_converter.WEEK_DAYS_ENGLISH[weekday_index],
            'is_weekend': weekday_index >= 5
        })
    
    return jsonify({
        'calendar': calendar_data,
        'timesheet_data': timesheet_data,
        'projects': projects,
        'month_days': month_days
    })

@app.route('/api/timesheet/save', methods=['POST'])
@login_required
def save_timesheet():
    """Save timesheet data"""
    user_session = get_current_user()
    
    try:
        data = request.get_json()
        
        # Extract data
        year = data.get('year')
        month = data.get('month')
        project_hours = data.get('project_hours', {})
        leave_hours = data.get('leave_hours', {})
        
        # Validate data
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        # Get month days for validation
        month_days = ethiopian_converter.get_ethiopian_month_days(year, month)
        
        # Initialize timesheet data
        timesheet_key = get_user_timesheet_key(user_session.user_id, year, month)
        timesheet_data = initialize_user_timesheet(user_session.user_id, year, month, month_days)
        
        # Save project hours
        for project_id, hours_data in project_hours.items():
            if project_id not in timesheet_data['projects']:
                timesheet_data['projects'][project_id] = {}
            
            for day, hours in hours_data.items():
                day_num = int(day)
                if 1 <= day_num <= month_days:
                    timesheet_data['projects'][project_id][day_num] = float(hours)
        
        # Save leave hours
        for leave_type, hours_data in leave_hours.items():
            if leave_type in timesheet_data['leave_entries']:
                for day, hours in hours_data.items():
                    day_num = int(day)
                    if 1 <= day_num <= month_days:
                        timesheet_data['leave_entries'][leave_type][day_num] = float(hours)
        
        # Update totals
        update_all_totals(user_session.user_id, year, month)
        
        return jsonify({
            'success': True,
            'message': 'Timesheet saved successfully',
            'totals': {
                'daily_totals': timesheet_data['daily_totals'],
                'leave_totals': timesheet_data['leave_totals'],
                'grand_totals': timesheet_data['grand_totals']
            }
        })
        
    except Exception as e:
        logger.error(f"Error saving timesheet: {e}")
        return jsonify({'error': 'Failed to save timesheet'}), 500

def update_all_totals(user_id, year, month):
    """Update all totals based on saved hours"""
    timesheet_key = get_user_timesheet_key(user_id, year, month)
    if timesheet_key not in timesheet_storage:
        return
    
    timesheet_data = timesheet_storage[timesheet_key]
    month_days = ethiopian_converter.get_ethiopian_month_days(year, month)
    
    # Reset totals
    timesheet_data['daily_totals'] = {day: 0.0 for day in range(1, month_days + 1)}
    timesheet_data['leave_totals'] = {day: 0.0 for day in range(1, month_days + 1)}
    timesheet_data['grand_totals'] = {day: 0.0 for day in range(1, month_days + 1)}
    
    # Calculate project totals
    for project_id, project_data in timesheet_data['projects'].items():
        total_hours = sum(project_data.values())
        
        # Update daily totals
        for day, hours in project_data.items():
            if 1 <= day <= month_days:
                timesheet_data['daily_totals'][day] += hours
        
        # Update project total in user_projects
        if user_id in user_projects and timesheet_key in user_projects[user_id]:
            for project in user_projects[user_id][timesheet_key]:
                if str(project['id']) == str(project_id):
                    project['total_hours'] = total_hours
                    project['hours'] = project_data
                    break
    
    # Calculate leave totals
    for leave_type, leave_data in timesheet_data['leave_entries'].items():
        for day, hours in leave_data.items():
            if 1 <= day <= month_days:
                timesheet_data['leave_totals'][day] += hours
    
    # Calculate grand totals
    for day in range(1, month_days + 1):
        timesheet_data['grand_totals'][day] = timesheet_data['daily_totals'][day] + timesheet_data['leave_totals'][day]

@app.route('/api/projects', methods=['GET', 'POST', 'DELETE'])
@login_required
def manage_projects():
    """Manage projects for timesheet"""
    user_session = get_current_user()
    
    if request.method == 'GET':
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        projects = initialize_user_projects(user_session.user_id, year, month)
        return jsonify({'projects': projects})
    
    elif request.method == 'POST':
        data = request.get_json()
        project_name = data.get('name', '').strip()
        allocated_hours = float(data.get('allocated_hours', 0))
        year = data.get('year')
        month = data.get('month')
        
        if not project_name:
            return jsonify({'error': 'Project name is required'}), 400
        
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        projects = initialize_user_projects(user_session.user_id, year, month)
        
        # Generate new project ID
        new_id = max([p['id'] for p in projects], default=0) + 1
        
        new_project = {
            'id': new_id,
            'name': project_name,
            'allocated_hours': allocated_hours,
            'total_hours': 0.0,
            'hours': {}
        }
        
        projects.append(new_project)
        
        return jsonify({
            'success': True,
            'project': new_project,
            'message': 'Project added successfully'
        })
    
    elif request.method == 'DELETE':
        project_id = request.args.get('project_id')
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not project_id:
            return jsonify({'error': 'Project ID is required'}), 400
        
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        timesheet_key = get_user_timesheet_key(user_session.user_id, year, month)
        
        if user_session.user_id in user_projects and timesheet_key in user_projects[user_session.user_id]:
            user_projects[user_session.user_id][timesheet_key] = [
                p for p in user_projects[user_session.user_id][timesheet_key] 
                if str(p['id']) != str(project_id)
            ]
        
        return jsonify({
            'success': True,
            'message': 'Project deleted successfully'
        })

@app.route('/api/timesheet/prefill', methods=['POST'])
@login_required
def prefill_timesheet():
    """Prefill default hours based on weekdays"""
    user_session = get_current_user()
    
    try:
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        month_days = ethiopian_converter.get_ethiopian_month_days(year, month)
        timesheet_key = get_user_timesheet_key(user_session.user_id, year, month)
        timesheet_data = initialize_user_timesheet(user_session.user_id, year, month, month_days)
        
        # Get user's first project (MERQ Internal)
        projects = initialize_user_projects(user_session.user_id, year, month)
        if not projects:
            return jsonify({'error': 'No projects found'}), 400
        
        first_project_id = str(projects[0]['id'])
        
        # Prefill hours based on weekdays
        prefilled_data = {}
        for day in range(1, month_days + 1):
            weekday_index = ethiopian_converter.get_ethiopian_weekday(year, month, day)
            
            # Set default hours based on weekday
            if weekday_index < 4:  # Monday-Thursday: 8 hours
                default_hours = 8.0
            elif weekday_index == 4:  # Friday: 8 hours
                default_hours = 8.0
            elif weekday_index == 5:  # Saturday: 4 hours
                default_hours = 4.0
            else:  # Sunday: 0 hours
                default_hours = 0.0
            
            # Set hours for first project
            if first_project_id not in timesheet_data['projects']:
                timesheet_data['projects'][first_project_id] = {}
            timesheet_data['projects'][first_project_id][day] = default_hours
            prefilled_data[day] = default_hours
        
        # Update project totals
        update_all_totals(user_session.user_id, year, month)
        
        return jsonify({
            'success': True,
            'message': 'Default hours prefilled successfully',
            'prefilled_data': prefilled_data,
            'totals': {
                'daily_totals': timesheet_data['daily_totals'],
                'leave_totals': timesheet_data['leave_totals'],
                'grand_totals': timesheet_data['grand_totals']
            }
        })
        
    except Exception as e:
        logger.error(f"Error prefilling timesheet: {e}")
        return jsonify({'error': 'Failed to prefill timesheet'}), 500

# Add this route to webapp.py (around line 400, after the prefill route)

@app.route('/api/timesheet/preview', methods=['POST'])
@login_required
def preview_timesheet_data():
    """Get timesheet preview data - FIXED VERSION"""
    user_session = get_current_user()
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        # Calculate totals for preview
        preview_data = calculate_timesheet_totals(user_session.user_id, year, month)
        preview_data.update({
            'year': year,
            'month': month,
            'month_name': ethiopian_converter.MONTHS_AMHARIC[month-1],
        })
        
        return jsonify({
            'success': True,
            'preview_data': preview_data
        })
        
    except Exception as e:
        logger.error(f"Error generating preview: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': f'Failed to generate preview: {str(e)}'}), 500
    
    
@app.route('/api/debug', methods=['GET'])
@login_required
def debug_info():
    """Debug endpoint to check application state"""
    user_session = get_current_user()
    
    debug_info = {
        'user_id': user_session.user_id,
        'user_name': user_session.full_name,
        'timesheet_storage_keys': list(timesheet_storage.keys()),
        'user_projects_keys': list(user_projects.keys()),
        'session_keys': list(session.keys())
    }
    
    return jsonify(debug_info)    

def calculate_timesheet_totals(user_id, year, month):
    """Calculate timesheet totals similar to desktop app"""
    timesheet_key = get_user_timesheet_key(user_id, year, month)
    if timesheet_key not in timesheet_storage:
        return {
            'project_totals': [],
            'daily_totals': {},
            'leave_totals': {},
            'grand_totals': {},
            'total_work_hours': 0,
            'total_leave_hours': 0,
            'grand_total': 0
        }
    
    timesheet_data = timesheet_storage[timesheet_key]
    month_days = ethiopian_converter.get_ethiopian_month_days(year, month)
    projects = user_projects.get(user_id, {}).get(timesheet_key, [])
    
    # Calculate project totals with percentages
    project_totals = []
    total_work_hours = sum(timesheet_data['daily_totals'].values())
    total_leave_hours = sum(timesheet_data['leave_totals'].values())
    grand_total = total_work_hours + total_leave_hours
    
    for project in projects:
        project_total = project.get('total_hours', 0.0)
        allocated_hours = project.get('allocated_hours', 0.0)
        
        project_totals.append({
            'name': project['name'],
            'total_hours': project_total,
            'allocated_hours': allocated_hours,
            'equiv_days': project_total / 8 if project_total > 0 else 0,
            'percent_direct': (project_total / total_work_hours * 100) if total_work_hours > 0 else 0,
            'percent_total': (project_total / (month_days * 8) * 100) if month_days * 8 > 0 else 0
        })
    
    return {
        'project_totals': project_totals,
        'daily_totals': timesheet_data['daily_totals'],
        'leave_totals': timesheet_data['leave_totals'],
        'grand_totals': timesheet_data['grand_totals'],
        'total_work_hours': total_work_hours,
        'total_leave_hours': total_leave_hours,
        'grand_total': grand_total,
        'year': year,
        'month': month,
        'month_name': ethiopian_converter.MONTHS_AMHARIC[month-1]
    }

@app.route('/timesheet/export')
@login_required
def export_timesheet():
    """Export timesheet to Excel using the template"""
    user_session = get_current_user()
    
    # Get parameters
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        flash('Year and month are required', 'error')
        return redirect(url_for('timesheet'))
    
    try:
        # Calculate totals for export
        totals_data = calculate_timesheet_totals(user_session.user_id, year, month)
        
        # Use the template file
        #template_file = os.path.join('templates', 'MERQ_TIMESHEET_ETH-CAL_TEMPLATE.xlsx')
        template_file = "MERQ_TIMESHEET_ETH-CAL_TEMPLATE.xlsx"
        if not os.path.exists(template_file):
            # Fallback to simple export
            return simple_export_timesheet(user_session, year, month, totals_data)
        
        # Load the template workbook
        from openpyxl import load_workbook
        workbook = load_workbook(template_file)
        worksheet = workbook.active
        
        # Safe method to update cells (handles merged cells)
        def safe_cell_update(cell_ref, value):
            """Safely update a cell, handling merged cells"""
            try:
                # Check if cell is part of a merged range
                cell = worksheet[cell_ref]
                for merged_range in list(worksheet.merged_cells.ranges):
                    if cell.coordinate in merged_range:
                        # Unmerge the range first
                        worksheet.unmerge_cells(str(merged_range))
                        break
                worksheet[cell_ref] = value
                return True
            except Exception as e:
                print(f"Warning: Could not update cell {cell_ref}: {e}")
                return False
        
        # Update header content - use safe method
        month_name = ethiopian_converter.MONTHS_AMHARIC[month-1]
        header_updates = [
            ('AJ19', f"{month_name} {year}"),
            ('C25', f"{month_name} {year}"), 
            ('AJ3', f"{month_name} {year}"),
            ('H5', user_session.full_name),
            ('X4', month_name),
            ('X5', year)
        ]
        
        for cell_ref, value in header_updates:
            safe_cell_update(cell_ref, value)
        
        # Fill project hours
        timesheet_key = get_user_timesheet_key(user_session.user_id, year, month)
        timesheet_data = timesheet_storage.get(timesheet_key, {})
        projects = user_projects.get(user_session.user_id, {}).get(timesheet_key, [])
        
        # Fill project data (rows 8-14)
        for i, project in enumerate(projects[:7]):  # Template has space for 7 projects
            row = 8 + i
            project_id = str(project['id'])
            project_hours = timesheet_data.get('projects', {}).get(project_id, {})
            
            # Update project name
            safe_cell_update(f'C{row}', project['name'])
            
            # Fill daily hours
            for day in range(1, 32):  # Up to 31 days
                if day in project_hours:
                    col = get_column_letter(3 + day)  # Starting from column D
                    hours = project_hours[day]
                    if hours > 0:
                        safe_cell_update(f'{col}{row}', hours)
        
        # Fill leave data (rows 16-21)
        leave_types = [
            ("vacation", 16),
            ("sick_leave", 17),
            ("holiday", 18),
            ("personal_leave", 19),
            ("bereavement", 20),
            ("other", 21)
        ]
        
        for leave_key, row in leave_types:
            leave_hours = timesheet_data.get('leave_entries', {}).get(leave_key, {})
            for day in range(1, 32):
                if day in leave_hours:
                    col = get_column_letter(3 + day)
                    hours = leave_hours[day]
                    if hours > 0:
                        safe_cell_update(f'{col}{row}', hours)
        
        # Update signature section with Ethiopian date
        eth_year, eth_month, eth_day = ethiopian_converter.gregorian_to_ethiopian(datetime.now())
        eth_date_str = f"{eth_day:02d}/{eth_month:02d}/{eth_year}"
        
        safe_cell_update('K29', eth_date_str)  # Employee date
        safe_cell_update('AJ29', eth_date_str)  # Supervisor date
        safe_cell_update('B29', user_session.full_name)
        
        # Update supervisor information if available
        if user_session.supervisor_name:
            safe_cell_update('P29', user_session.supervisor_name)
        if user_session.supervisor_position_title:
            safe_cell_update('T29', user_session.supervisor_position_title)

        # Add timestamp to filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            workbook.save(tmp_file.name)
            
            # Return file
            filename = f"MERQ_TIMESHEET_{user_session.full_name}_{month_name}_{year}_{timestamp}.xlsx"
            return send_file(
                tmp_file.name,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
    except Exception as e:
        logger.error(f"Error exporting timesheet with template: {e}")
        # Fallback to simple export
        return simple_export_timesheet(user_session, year, month, calculate_timesheet_totals(user_session.user_id, year, month))

def simple_export_timesheet(user_session, year, month, totals_data):
    """Simple export fallback"""
    try:
        # Create Excel file
        output = BytesIO()
        
        # Create summary data
        data = {
            'Employee Name': [user_session.full_name],
            'Month': [ethiopian_converter.MONTHS_AMHARIC[month-1]],
            'Year': [year],
            'Total Work Hours': [totals_data['total_work_hours']],
            'Total Leave Hours': [totals_data['total_leave_hours']],
            'Grand Total': [totals_data['grand_total']]
        }
        
        df = pd.DataFrame(data)
        
        # Create Excel file
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Timesheet Summary', index=False)
            
            # Add project details
            project_data = []
            for project in totals_data['project_totals']:
                project_data.append({
                    'Project Name': project['name'],
                    'Total Hours': project['total_hours'],
                    'Allocated Hours': project['allocated_hours'],
                    'Equivalent Days': project['equiv_days']
                })
            
            if project_data:
                project_df = pd.DataFrame(project_data)
                project_df.to_excel(writer, sheet_name='Project Details', index=False)
        
        output.seek(0)
        
        # Return file
        filename = f"MERQ_Timesheet_{user_session.full_name}_{year}_{month}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error in simple export: {e}")
        flash('Error exporting timesheet', 'error')
        return redirect(url_for('timesheet'))

@app.route('/timesheet/submit', methods=['POST'])
@login_required
def submit_timesheet():
    """Submit timesheet to HR"""
    user_session = get_current_user()
    
    try:
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        # Calculate totals for submission
        totals_data = calculate_timesheet_totals(user_session.user_id, year, month)
        
        # Create submission record
        submission_data = {
            'user_id': user_session.user_id,
            'user_name': user_session.full_name,
            'year': year,
            'month': month,
            'submission_date': datetime.now().isoformat(),
            'totals': totals_data
        }
        
        logger.info(f"Timesheet submitted: {submission_data}")
        
        return jsonify({
            'success': True,
            'message': 'Timesheet submitted to HR successfully',
            'submission_data': submission_data
        })
        
    except Exception as e:
        logger.error(f"Error submitting timesheet: {e}")
        return jsonify({'error': 'Failed to submit timesheet'}), 500

@app.route('/timesheet/clear', methods=['POST'])
@login_required
def clear_timesheet():
    """Clear timesheet data"""
    user_session = get_current_user()
    
    try:
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return jsonify({'error': 'Year and month are required'}), 400
        
        timesheet_key = get_user_timesheet_key(user_session.user_id, year, month)
        
        if timesheet_key in timesheet_storage:
            # Clear projects data but keep structure
            timesheet_storage[timesheet_key]['projects'] = {}
            
            # Clear leave entries
            for leave_type in timesheet_storage[timesheet_key]['leave_entries']:
                timesheet_storage[timesheet_key]['leave_entries'][leave_type] = {
                    day: 0.0 for day in timesheet_storage[timesheet_key]['leave_entries'][leave_type]
                }
        
        # Reset project totals
        if user_session.user_id in user_projects and timesheet_key in user_projects[user_session.user_id]:
            for project in user_projects[user_session.user_id][timesheet_key]:
                project['total_hours'] = 0.0
                project['hours'] = {}
        
        # Update totals
        update_all_totals(user_session.user_id, year, month)
        
        return jsonify({
            'success': True,
            'message': 'Timesheet cleared successfully',
            'totals': {
                'daily_totals': timesheet_storage[timesheet_key]['daily_totals'],
                'leave_totals': timesheet_storage[timesheet_key]['leave_totals'],
                'grand_totals': timesheet_storage[timesheet_key]['grand_totals']
            }
        })
        
    except Exception as e:
        logger.error(f"Error clearing timesheet: {e}")
        return jsonify({'error': 'Failed to clear timesheet'}), 500

@app.route('/profile')
@login_required
def profile():
    """User profile page"""
    user_session = get_current_user()
    return render_template('profile.html', user=user_session)

@app.route('/help')
@login_required
def help_page():
    """Help and instructions page"""
    user_session = get_current_user()
    return render_template('help.html', user=user_session)

# Utility functions
def normalize_email(email):
    """Normalize email input (from timesheet.py LoginWindow)"""
    email = email.strip().lower()
    
    # Remove any existing @merqconsultancy.org if present to avoid duplication
    if email.endswith('@merqconsultancy.org'):
        email = email.replace('@merqconsultancy.org', '')
    
    # Add the domain
    return email + '@merqconsultancy.org'

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

if __name__ == '__main__':
    # Create templates and static directories if they don't exist
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)
    os.makedirs('static/images', exist_ok=True)
    
    # Run the application
    app.run(debug=True, host='0.0.0.0', port=5000)