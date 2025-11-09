# MERQ Timesheet Management System
# Version: 1.0.0.1
# Author: Michael Kifle Teferra
# Date: Novmeber 2025
# Description: Core functionalities for MERQ Timesheet application including
#              database management, user authentication, email handling, excel formatting,
#              and Ethiopian date conversion.
# License: MIT License

# Required Libraries
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import pandas as pd
from datetime import datetime, timedelta
import sys
import json
import tempfile
import subprocess
from packaging import version
import os
from PIL import Image, ImageTk, ImageDraw, ImageFont
import math
import requests
import json
from typing import Tuple, Optional, Dict, List
import threading
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import webbrowser
import sqlite3
import hashlib
import bcrypt
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Required packages: pip install pandas openpyxl pillow requests

class DatabaseManager:
    """Manages SQLite database operations"""
    
    def __init__(self, db_path="./merq_timesheet_db.sqlite"):
        self.db_path = db_path
        self.init_database()
    
    def init_database(self):
        """Initialize database connection and verify structure"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Verify tables exist
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = cursor.fetchall()
            table_names = [table[0] for table in tables]
            
            required_tables = ['users', 'departments', 'positions']
            for table in required_tables:
                if table not in table_names:
                    raise Exception(f"Required table '{table}' not found in database")
            
            conn.close()
            return True
        except Exception as e:
            print(f"Database initialization error: {e}")
            return False
    
    def get_user_by_email(self, email):
        """Get user by email address"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT u.*, 
                    p.position_title, 
                    d.department_name, 
                    sup.full_name AS supervisor_name, 
                    sup.email AS supervisor_email, 
                    sup_pos.position_title AS supervisor_position_title
                FROM users u
                LEFT JOIN positions p ON u.position_id = p.position_id
                LEFT JOIN departments d ON u.department_id = d.department_id
                LEFT JOIN users sup ON u.supervisor_id = sup.user_id
                LEFT JOIN positions sup_pos ON sup.position_id = sup_pos.position_id
                WHERE u.email = ? AND u.is_active = 1
            """, (email,))
            
            user = cursor.fetchone()
            conn.close()
            
            if user:
                # Convert to dictionary with column names
                columns = [description[0] for description in cursor.description]
                return dict(zip(columns, user))
            return None
        except Exception as e:
            print(f"Error getting user by email: {e}")
            return None
    
    def get_hr_users(self):
        """Get all HR users (position_id = 15)"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT u.*, p.position_title, d.department_name
                FROM users u
                JOIN positions p ON u.position_id = p.position_id
                JOIN departments d ON u.department_id = d.department_id
                WHERE u.position_id = 15 AND u.is_active = 1
            """)
            
            hr_users = cursor.fetchall()
            conn.close()
            
            if hr_users:
                columns = [description[0] for description in cursor.description]
                return [dict(zip(columns, user)) for user in hr_users]
            return []
        except Exception as e:
            print(f"Error getting HR users: {e}")
            return []
    
    def validate_user_credentials(self, email, password):
        """Validate user credentials using bcrypt with email normalization"""
        try:
            # Normalize email for query
            normalized_email = self.normalize_email_for_query(email)
            
            user = self.get_user_by_email(normalized_email)
            if not user:
                return None

            stored_hash = user.get('password_hash')
            if not stored_hash:
                return None

            # Stored hash might be bytes or str. Normalize to str.
            if isinstance(stored_hash, bytes):
                stored_hash = stored_hash.decode('utf-8')

            # PHP uses $2y$ prefix; Python bcrypt accepts $2b$.
            # Normalizing $2y$ -> $2b$ is safe for verification.
            if stored_hash.startswith("$2y$"):
                stored_hash = "$2b$" + stored_hash[4:]

            # bcrypt.checkpw expects bytes
            password_bytes = password.encode('utf-8')
            stored_bytes = stored_hash.encode('utf-8')

            if bcrypt.checkpw(password_bytes, stored_bytes):
                return user

            return None
        except Exception as e:
            print(f"Error validating credentials: {e}")
            return None
    
    def normalize_email_for_query(self, email):
        """Normalize email for database query"""
        email = email.strip().lower()
        
        # Ensure it has the merqconsultancy.org domain
        if '@' not in email:
            email = email + '@merqconsultancy.org'
        elif not email.endswith('@merqconsultancy.org'):
            # Extract username and add correct domain
            username = email.split('@')[0]
            email = username + '@merqconsultancy.org'
            
        return email

class UserSession:
    """Manages user session data"""
    
    def __init__(self, user_data):
        self.user_data = user_data
        self.logged_in = True
        self.timesheet_data = {}
    
    @property
    def full_name(self):
        return self.user_data.get('full_name', '')
    
    @property
    def email(self):
        return self.user_data.get('email', '')
    
    @property
    def position(self):
        return self.user_data.get('position_title', '')
    
    @property
    def department(self):
        return self.user_data.get('department_name', '')
    
    @property
    def supervisor_name(self):
        return self.user_data.get('supervisor_name', '')

    @property
    def supervisor_position_title(self):
        return self.user_data.get('supervisor_position_title', '')
    
    @property
    def supervisor_email(self):
        return self.user_data.get('supervisor_email', '')
    
    @property
    def employee_id(self):
        return self.user_data.get('employee_id', '')
    
    @property
    def user_id(self):
        return self.user_data.get('user_id', '')

class LoginWindow:
    """Handles user authentication with modern design"""
    
    def __init__(self, parent, db_manager, on_login_success):
        self.parent = parent
        self.db_manager = db_manager
        self.on_login_success = on_login_success
        self.create_modern_login_window()
    
    def create_modern_login_window(self):
        """Create modern and attractive login window"""
        self.login_window = tk.Toplevel(self.parent)
        self.login_window.title("MERQ Timesheet - Login")
        self.login_window.geometry("970x900")
        self.login_window.resizable(True, True)
        self.login_window.configure(bg=ModernTheme.BACKGROUND)
        self.login_window.transient(self.parent)
        self.login_window.grab_set()
        
        # Center window
        self.login_window.update_idletasks()
        x = (self.login_window.winfo_screenwidth() // 2) - (970 // 2)
        y = (self.login_window.winfo_screenheight() // 2) - (900 // 2)
        self.login_window.geometry(f"970x900+{x}+{y}")
        
        # Main container with gradient effect simulation
        main_container = ttk.Frame(self.login_window, style="Modern.TFrame")
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header section with logo
        header_frame = ttk.Frame(main_container, style="Card.TFrame", height=120)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        header_frame.pack_propagate(False)
        
        # Logo and title
        try:
            if os.path.exists("merq.png"):
                img = Image.open("merq.png")
                img = img.resize((80, 80), Image.Resampling.LANCZOS)
                self.login_logo_img = ImageTk.PhotoImage(img)
                logo_label = ttk.Label(header_frame, image=self.login_logo_img, 
                                     background=ModernTheme.CARD_BG)
                logo_label.pack(side=tk.LEFT, padx=30, pady=20)
        except:
            pass
        
        title_frame = ttk.Frame(header_frame, style="Card.TFrame")
        title_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=20)
        
        ttk.Label(title_frame, text="MERQ CONSULTANCY", 
                 style="Title.TLabel", font=("Arial", 20, "bold")).pack(anchor=tk.W)
        ttk.Label(title_frame, text="ወርሃዊ የስራ ሰዓት መከታተያ / Monthly Timesheet Tracker", 
                 style="Subtitle.TLabel", font=("Arial", 12)).pack(anchor=tk.W)
        ttk.Label(title_frame, text="Secure Employee Login", 
                 style="Subtitle.TLabel", font=("Arial", 10), 
                 foreground=ModernTheme.SUCCESS).pack(anchor=tk.W)
        
        # Login form container
        form_container = ttk.Frame(main_container, style="Card.TFrame", padding=40)
        form_container.pack(fill=tk.BOTH, expand=True)
        
        # Welcome message
        welcome_frame = ttk.Frame(form_container, style="Card.TFrame")
        welcome_frame.pack(fill=tk.X, pady=(0, 30))
        
        ttk.Label(welcome_frame, text="🔐 Welcome Back!", 
                 style="Title.TLabel", font=("Arial", 18, "bold")).pack(anchor=tk.W)
        ttk.Label(welcome_frame, text="Sign in to your MERQ Timesheet account", 
                 style="Subtitle.TLabel", font=("Arial", 11)).pack(anchor=tk.W)
        
        # Email frame with modern styling
        email_frame = ttk.LabelFrame(form_container, text="📧 Email Address / ኢሜይል አድራሻ", 
                                   style="Card.TFrame", padding=15)
        email_frame.pack(fill=tk.X, pady=15)
        
        ttk.Label(email_frame, text="Enter your MERQ email:", style="Dark.TLabel", 
                 font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 8))
        
        email_instructions = ttk.Label(email_frame, 
                                     text="You can enter: username, username@merqconsultancy.org, or full email",
                                     style="Dark.TLabel", font=("Arial", 8),
                                     foreground=ModernTheme.TEXT_MUTED)
        email_instructions.pack(anchor=tk.W, pady=(0, 5))
        
        self.email_var = tk.StringVar()
        email_entry = ttk.Entry(email_frame, textvariable=self.email_var, 
                               font=("Arial", 12), style="Modern.TEntry",
                               width=30)
        email_entry.pack(fill=tk.X, pady=(5, 0))
        email_entry.focus()
        
        # Auto-complete domain
        def auto_complete_domain(event=None):
            email = self.email_var.get().strip()
            if email and '@' not in email:
                if not email.endswith('@merqconsultancy.org'):
                    self.email_var.set(email + '@merqconsultancy.org')
        
        email_entry.bind('<FocusOut>', auto_complete_domain)
        
        # Password frame
        password_frame = ttk.LabelFrame(form_container, text="🔒 Password / የይለፍ ቃል", 
                                      style="Card.TFrame", padding=15)
        password_frame.pack(fill=tk.X, pady=15)
        
        ttk.Label(password_frame, text="Enter your password:", style="Dark.TLabel", 
                 font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 8))
        
        self.password_var = tk.StringVar()
        password_entry = ttk.Entry(password_frame, textvariable=self.password_var, 
                                  font=("Arial", 12), style="Modern.TEntry",
                                  show="•", width=30)
        password_entry.pack(fill=tk.X, pady=(5, 0))
        
        # Bind Enter key to login
        email_entry.bind('<Return>', lambda e: password_entry.focus())
        password_entry.bind('<Return>', lambda e: self.login())
        
        # Login button with modern styling
        login_btn_frame = ttk.Frame(form_container, style="Card.TFrame")
        login_btn_frame.pack(fill=tk.X, pady=30)
        
        login_btn = ttk.Button(login_btn_frame, text="🚀 Login / ግባ", 
                              command=self.login, style="Primary.TButton",
                              width=20)
        login_btn.pack(pady=10)
        
        # Status label with better visibility
        self.status_label = ttk.Label(login_btn_frame, text="", style="Dark.TLabel", 
                                     font=("Arial", 10), foreground=ModernTheme.ACCENT)
        self.status_label.pack()
        
        # Instructions section
        instructions_frame = ttk.LabelFrame(form_container, text="💡 Login Instructions", 
                                          style="Card.TFrame", padding=15)
        instructions_frame.pack(fill=tk.X, pady=20)
        
        instructions_text = """
        How to login:
        • Use your MERQ email credentials
        • You can enter just your username (e.g., 'michaelktd'), username with domain (e.g., 'michaelktd@merqconsultancy.org'), or full email
        • The system will automatically add @merqconsultancy.org if you only enter username
        • Contact IT support if you forgot your password
        
        እንዴት መግባት እንደሚቻል:
        • የመርቅ ኢሜይል መለያዎትን ይጠቀሙ
        • የተጠቃሚ ስምዎን ብቻ (ለምሳሌ 'michaelktd') ማስገባት ትችላላችሁ
        • ስለማስገባት ችግር ካጋጠመዎት ከአይቲ ክፍል ይጠይቁ
        """
        
        instructions_widget = scrolledtext.ScrolledText(instructions_frame, wrap=tk.WORD, 
                                                      font=("Arial", 9), height=6,
                                                      bg=ModernTheme.CARD_BG, relief="flat")
        instructions_widget.pack(fill=tk.BOTH, expand=True)
        instructions_widget.insert(tk.END, instructions_text)
        instructions_widget.config(state=tk.DISABLED)
        
        # Footer
        footer_frame = ttk.Frame(main_container, style="Card.TFrame", height=30)
        footer_frame.pack(fill=tk.X, pady=(20, 0))
        footer_frame.pack_propagate(False)
        
        ttk.Label(footer_frame, text="🔒 Secure Login | MERQ Consultancy PLC 2025", 
                 style="Dark.TLabel", font=("Arial", 8),
                 foreground=ModernTheme.TEXT_MUTED).pack(pady=5)
    
    def normalize_email(self, email):
        """Normalize email input to handle various formats"""
        email = email.strip().lower()
        
        # Remove any existing @merqconsultancy.org if present to avoid duplication
        if email.endswith('@merqconsultancy.org'):
            email = email.replace('@merqconsultancy.org', '')
        
        # Add the domain
        return email + '@merqconsultancy.org'
    
    def login(self):
        """Handle login attempt with improved email handling"""
        email_input = self.email_var.get().strip()
        password = self.password_var.get()
        
        if not email_input or not password:
            self.status_label.config(text="Please enter both email and password", 
                                   foreground=ModernTheme.ACCENT)
            return
        
        # Normalize email
        normalized_email = self.normalize_email(email_input)
        
        if not re.match(r'^[^@]+@[^@]+\.[^@]+$', normalized_email):
            self.status_label.config(text="Please enter a valid email address",
                                   foreground=ModernTheme.ACCENT)
            return
        
        # Update status
        self.status_label.config(text="Authenticating... Please wait", 
                               foreground=ModernTheme.INFO)
        self.login_window.update()
        
        # Validate credentials
        user = self.db_manager.validate_user_credentials(normalized_email, password)
        
        if user:
            self.status_label.config(text="✅ Login successful! Loading...", 
                                   foreground=ModernTheme.SUCCESS)
            self.login_window.after(1000, lambda: self.finalize_login(user))
        else:
            self.status_label.config(text="❌ Invalid email or password",
                                   foreground=ModernTheme.ACCENT)
    
    def finalize_login(self, user):
        """Finalize login process"""
        self.login_window.destroy()
        self.on_login_success(user)

class EmailManager:
    """Handles email sending functionality"""
    
    def __init__(self):
        self.smtp_server = "smtp.gmail.com"  # Change based on your email provider
        self.smtp_port = 587
    
    def send_timesheet_email(self, timesheet_file, user_session, hr_email="haymanot.a@merqconsultancy.org"):
        """Send timesheet via email to HR"""
        try:
            # Create message
            msg = MIMEMultipart()
            msg['From'] = user_session.email
            msg['To'] = hr_email
            msg['Subject'] = f"Timesheet Submission - {user_session.full_name} - {datetime.now().strftime('%B %Y')}"
            
            # Email body
            body = f"""
            Dear HR Department,
            
            Please find attached my timesheet for your review and approval.
            
            Employee Details:
            - Name: {user_session.full_name}
            - Position: {user_session.position}
            - Department: {user_session.department}
            - Employee ID: {user_session.employee_id}
            - Supervisor: {user_session.supervisor_name}
            - Supervisor Position: {user_session.supervisor_position_title}
            
            This timesheet has been generated through the MERQ Timesheet System.
            
            Best regards,
            {user_session.full_name}
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach timesheet file
            with open(timesheet_file, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            filename = os.path.basename(timesheet_file)
            part.add_header('Content-Disposition', f"attachment; filename= {filename}")
            msg.attach(part)
            
            # Send email (you'll need to configure SMTP settings)
            # This is a template - you'll need to add your email credentials
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            # server.login('your_email@merqconsultancy.org', 'your_password')  # Configure this
            text = msg.as_string()
            # server.sendmail(user_session.email, hr_email, text)  # Uncomment when configured
            server.quit()
            
            return True
        except Exception as e:
            print(f"Email sending error: {e}")
            return False

class EthiopianDateConverter:
    """Ethiopian date conversion utilities using accurate algorithms"""
    
    # Ethiopian month names
    MONTHS_AMHARIC = [
        "መስከረም", "ጥቅምት", "ኅዳር", "ታኅሣሥ", 
        "ጥር", "የካቲት", "መጋቢት", "ሚያዝያ", 
        "ግንቦት", "ሰኔ", "ሐምሌ", "ነሐሴ", "ጳጉሜ"
    ]
    
    MONTHS_ENGLISH = [
        "Meskerem", "Tikimt", "Hidar", "Tahsas",
        "Tir", "Yekatit", "Megabit", "Miyazya",
        "Ginbot", "Sene", "Hamle", "Nehase", "Pagume"
    ]
    
    # Week days in Amharic
    WEEK_DAYS_AMHARIC = ["ሰኞ", "ማክሰኞ", "ረቡዕ", "ሐሙስ", "ዓርብ", "ቅዳሜ", "እሁድ"]
    WEEK_DAYS_ENGLISH = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    
    @staticmethod
    def gregorian_to_ethiopian(greg_date: datetime) -> Tuple[int, int, int]:
        """
        Convert Gregorian date to Ethiopian date using accurate algorithm
        Based on the Ethiopian calendar system
        """
        greg_year = greg_date.year
        greg_month = greg_date.month
        greg_day = greg_date.day
        
        # Ethiopian New Year in Gregorian calendar is September 11 or 12
        # Determine if it's before or after Ethiopian New Year
        if (greg_month > 9) or (greg_month == 9 and greg_day >= 12):
            eth_year = greg_year - 7
        else:
            eth_year = greg_year - 8
        
        # Calculate days from Ethiopian New Year (Meskerem 1)
        if eth_year % 4 == 3:  # Ethiopian leap year (Year of John)
            new_year_day = 12  # September 12 in Gregorian
        else:
            new_year_day = 11  # September 11 in Gregorian
        
        new_year = datetime(greg_year, 9, new_year_day)
        
        if greg_date >= new_year:
            # Current Gregorian year's Ethiopian New Year has passed
            days_diff = (greg_date - new_year).days
        else:
            # Use previous year's Ethiopian New Year
            prev_eth_year = eth_year - 1
            if prev_eth_year % 4 == 3:
                prev_new_year_day = 12
            else:
                prev_new_year_day = 11
            prev_new_year = datetime(greg_year - 1, 9, prev_new_year_day)
            days_diff = (greg_date - prev_new_year).days
        
        # Calculate Ethiopian month and day
        eth_month = (days_diff // 30) + 1
        eth_day = (days_diff % 30) + 1
        
        # Handle Pagume (13th month) - special case
        if eth_month == 13:
            # Ethiopian leap year has 6 days in Pagume, normal year has 5
            pagume_days = 6 if eth_year % 4 == 3 else 5
            
            if eth_day > pagume_days:
                eth_year += 1
                eth_month = 1
                eth_day = eth_day - pagume_days
        
        return eth_year, eth_month, eth_day
    
    @staticmethod
    def ethiopian_to_gregorian(eth_year: int, eth_month: int, eth_day: int) -> datetime:
        """
        Convert Ethiopian date to Gregorian date using accurate algorithm
        """
        # Calculate approximate Gregorian year
        greg_year = eth_year + 7 if (eth_month >= 1 and eth_month <= 5) or (eth_month == 13 and eth_day <= 5) else eth_year + 8
        
        # Calculate days from Ethiopian New Year
        days_from_new_year = (eth_month - 1) * 30 + (eth_day - 1)
        
        # Handle Pagume adjustments
        if eth_month == 13:
            if eth_year % 4 == 3:  # Leap year
                days_from_new_year = 360 + (eth_day - 1)
            else:
                days_from_new_year = 359 + (eth_day - 1)
        
        # Determine Ethiopian New Year in Gregorian calendar
        if eth_year % 4 == 3:  # Ethiopian leap year
            new_year_day = 12
        else:
            new_year_day = 11
        
        new_year = datetime(greg_year, 9, new_year_day)
        
        # Calculate Gregorian date
        greg_date = new_year + timedelta(days=days_from_new_year)
        return greg_date
    
    @staticmethod
    def get_ethiopian_month_days(year: int, month: int) -> int:
        """Get number of days in Ethiopian month"""
        if month == 13:  # Pagume
            if year % 4 == 3:  # Ethiopian leap year
                return 6
            else:
                return 5
        else:
            return 30
    
    @staticmethod
    def get_ethiopian_weekday(eth_year: int, eth_month: int, eth_day: int) -> int:
        """Get weekday for Ethiopian date (0=Monday, 6=Sunday)"""
        greg_date = EthiopianDateConverter.ethiopian_to_gregorian(eth_year, eth_month, eth_day)
        return greg_date.weekday()
    
    @staticmethod
    def format_ethiopian_date(eth_year: int, eth_month: int, eth_day: int) -> str:
        """Format Ethiopian date as dd/mm/yyyy"""
        return f"{eth_day}/{eth_month}/{eth_year}"
    
    @staticmethod
    def get_current_ethiopian_date() -> dict:
        """
        Get current Ethiopian date from API with fallback to local calculation
        """
        try:
            # Try to get from API first
            response = requests.get("https://api.ethioall.com/date/api", timeout=5)
            if response.status_code == 200:
                data = response.json()
                return {
                    'year': data.get('year', 2018),
                    'month': data.get('month_number', 2),
                    'day': data.get('date', 27),
                    'month_name': data.get('month_amharic', 'ጥቅምት'),
                    'weekday': data.get('day_amharic', 'ሐሙስ'),
                    'english_month': data.get('month_english', 'Tikimt'),
                    'english_weekday': data.get('day_english', 'Thursday')
                }
        except Exception as e:
            print(f"API call failed: {e}, using local calculation")
        
        # Fallback: Calculate locally with corrected algorithm
        current_greg = datetime.now()
        eth_year, eth_month, eth_day = EthiopianDateConverter.gregorian_to_ethiopian(current_greg)
        weekday_index = EthiopianDateConverter.get_ethiopian_weekday(eth_year, eth_month, eth_day)
        
        return {
            'year': eth_year,
            'month': eth_month,
            'day': eth_day,
            'month_name': EthiopianDateConverter.MONTHS_AMHARIC[eth_month - 1],
            'weekday': EthiopianDateConverter.WEEK_DAYS_AMHARIC[weekday_index],
            'english_month': EthiopianDateConverter.MONTHS_ENGLISH[eth_month - 1],
            'english_weekday': EthiopianDateConverter.WEEK_DAYS_ENGLISH[weekday_index]
        }

class ModernTheme:
    """Modern theme colors and styles with better contrast"""
    
    # Professional color scheme with good contrast
    PRIMARY = "#2C3E50"      # Dark blue-gray
    SECONDARY = "#3498DB"    # Bright blue
    ACCENT = "#E74C3C"       # Red
    SUCCESS = "#27AE60"      # Green
    WARNING = "#F39C12"      # Orange
    INFO = "#2980B9"         # Blue
    LIGHT = "#ECF0F1"        # Light gray
    DARK = "#2C3E50"         # Dark blue-gray
    BACKGROUND = "#F8F9FA"   # Off-white
    CARD_BG = "#FFFFFF"      # White
    CONTENT_BG = "#F8F9FA"   # Content background
    BORDER = "#BDC3C7"       # Gray border
    
    BUTTON_TEXT = "#0D2769"       # Navy Blue
    BUTTON_HOVER_TEXT = "#2980B9"   # Hover Blue
    
    # Excel colors matching the template
    EXCEL_HEADER_BG = "3498DB"      # Blue header
    EXCEL_PROJECT_BG = "FFF3CD"     # Light yellow for projects
    EXCEL_LEAVE_BG = "F8D7DA"       # Light red for leave
    EXCEL_TOTAL_BG = "2980B9"       # Dark blue for totals
    EXCEL_WEEKEND_BG = "D4EDDA"     # Light green for weekends
    
    # Text colors with good contrast
    TEXT_DARK = "#2C3E50"    # Dark text
    TEXT_LIGHT = "#F7FBFF"   # Light text
    TEXT_MUTED = "#7F8C8D"   # Muted text
    
    @staticmethod
    def configure_styles():
        """Configure ttk styles for modern look with good contrast"""
        style = ttk.Style()
        
        # Configure main styles
        style.configure("Modern.TFrame", background=ModernTheme.BACKGROUND)
        style.configure("Card.TFrame", background=ModernTheme.CARD_BG, relief="solid", borderwidth=1)
        
        # Label styles
        style.configure("Header.TLabel", background=ModernTheme.PRIMARY, foreground=ModernTheme.TEXT_LIGHT, 
                       font=("Arial", 14, "bold"))
        style.configure("Title.TLabel", background=ModernTheme.BACKGROUND, foreground=ModernTheme.TEXT_DARK, 
                       font=("Arial", 16, "bold"))
        style.configure("Subtitle.TLabel", background=ModernTheme.BACKGROUND, foreground=ModernTheme.SECONDARY, 
                       font=("Arial", 12))
        style.configure("Dark.TLabel", background=ModernTheme.CARD_BG, foreground=ModernTheme.TEXT_DARK)
        style.configure("Light.TLabel", background=ModernTheme.PRIMARY, foreground=ModernTheme.TEXT_LIGHT)
        
        # Button styles with good contrast
        style.configure("Primary.TButton", background=ModernTheme.SECONDARY, foreground=ModernTheme.BUTTON_TEXT, 
                       font=("Arial", 10, "bold"), focuscolor=ModernTheme.SECONDARY)
        style.configure("Success.TButton", background=ModernTheme.SUCCESS, foreground=ModernTheme.BUTTON_TEXT, 
                       font=("Arial", 10, "bold"), focuscolor=ModernTheme.SUCCESS)
        style.configure("Danger.TButton", background=ModernTheme.ACCENT, foreground=ModernTheme.BUTTON_TEXT, 
                       font=("Arial", 10, "bold"), focuscolor=ModernTheme.ACCENT)
        style.configure("Warning.TButton", background=ModernTheme.WARNING, foreground=ModernTheme.TEXT_DARK, 
                       font=("Arial", 10, "bold"), focuscolor=ModernTheme.WARNING)
        
        # Entry and Combobox styles
        style.configure("Modern.TEntry", fieldbackground="white", foreground=ModernTheme.TEXT_DARK, 
                       borderwidth=1, relief="solid")
        style.configure("Modern.TCombobox", fieldbackground="white", foreground=ModernTheme.TEXT_DARK, 
                       borderwidth=1, relief="solid")
        
        # Scrollbar styles
        style.configure("Modern.Vertical.TScrollbar", background=ModernTheme.LIGHT, 
                       darkcolor=ModernTheme.SECONDARY, lightcolor=ModernTheme.SECONDARY)
        style.configure("Modern.Horizontal.TScrollbar", background=ModernTheme.LIGHT, 
                       darkcolor=ModernTheme.SECONDARY, lightcolor=ModernTheme.SECONDARY)

class ExcelFormatter:
    """Handles Excel formatting and styling"""
    
    @staticmethod
    def apply_workbook_styles(workbook):
        """Apply consistent styles to the workbook"""
        # Header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF", size=12)
        header_style.fill = PatternFill(start_color=ModernTheme.EXCEL_HEADER_BG, end_color=ModernTheme.EXCEL_HEADER_BG, fill_type="solid")
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Project style
        project_style = NamedStyle(name="project_style")
        project_style.font = Font(bold=True, size=10)
        project_style.fill = PatternFill(start_color=ModernTheme.EXCEL_PROJECT_BG, end_color=ModernTheme.EXCEL_PROJECT_BG, fill_type="solid")
        project_style.alignment = Alignment(horizontal='center', vertical='center')
        project_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Leave style
        leave_style = NamedStyle(name="leave_style")
        leave_style.font = Font(bold=True, size=10)
        leave_style.fill = PatternFill(start_color=ModernTheme.EXCEL_LEAVE_BG, end_color=ModernTheme.EXCEL_LEAVE_BG, fill_type="solid")
        leave_style.alignment = Alignment(horizontal='center', vertical='center')
        leave_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Total style
        total_style = NamedStyle(name="total_style")
        total_style.font = Font(bold=True, color="FFFFFF", size=11)
        total_style.fill = PatternFill(start_color=ModernTheme.EXCEL_TOTAL_BG, end_color=ModernTheme.EXCEL_TOTAL_BG, fill_type="solid")
        total_style.alignment = Alignment(horizontal='center', vertical='center')
        total_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Weekend style
        weekend_style = NamedStyle(name="weekend_style")
        weekend_style.font = Font(size=9)
        weekend_style.fill = PatternFill(start_color=ModernTheme.EXCEL_WEEKEND_BG, end_color=ModernTheme.EXCEL_WEEKEND_BG, fill_type="solid")
        weekend_style.alignment = Alignment(horizontal='center', vertical='center')
        weekend_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Normal cell style
        normal_style = NamedStyle(name="normal_style")
        normal_style.font = Font(size=9)
        normal_style.alignment = Alignment(horizontal='center', vertical='center')
        normal_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add styles to workbook
        workbook.add_named_style(header_style)
        workbook.add_named_style(project_style)
        workbook.add_named_style(leave_style)
        workbook.add_named_style(total_style)
        workbook.add_named_style(weekend_style)
        workbook.add_named_style(normal_style)
        
        return {
            'header': header_style,
            'project': project_style,
            'leave': leave_style,
            'total': total_style,
            'weekend': weekend_style,
            'normal': normal_style
        }
    
    @staticmethod
    def create_watermark(workbook, worksheet, text="MERQ CONSULTANCY"):
        """Create a watermark on the worksheet"""
        try:
            # Create a temporary image for watermark
            img = Image.new('RGBA', (400, 200), (255, 255, 255, 0))
            draw = ImageDraw.Draw(img)
            
            # Try to use a font (this might not work on all systems)
            try:
                font = ImageFont.truetype("arial.ttf", 40)
            except:
                try:
                    font = ImageFont.truetype("/usr/share/fonts/truetype/freefont/FreeSans.ttf", 40)
                except:
                    font = ImageFont.load_default()
            
            # Draw semi-transparent text
            draw.text((50, 80), text, fill=(200, 200, 200, 100), font=font)
            
            # Save temporary image
            watermark_path = "watermark_temp.png"
            img.save(watermark_path)
            
            # Add to worksheet (this is a simplified approach)
            # In a real implementation, you might use openpyxl's header/footer or background image
            
            # Clean up
            if os.path.exists(watermark_path):
                os.remove(watermark_path)
                
        except Exception as e:
            print(f"Watermark creation failed: {e}")

class UpdateManager:
    """Handles application updates in background"""
    
    def __init__(self, app):
        self.app = app
        self.current_version = "1.0.0.0"
        self.update_url = "https://app.merqconsultancy.org/apps/timesheet/desktop/version.json"
        self.download_url = "https://app.merqconsultancy.org/apps/timesheet/desktop/"
    
    def check_for_updates_async(self):
        """Check for updates in background thread without blocking UI"""
        def update_check():
            try:
                # Add a small delay to ensure UI is loaded
                time.sleep(2)
                
                response = requests.get(self.update_url, timeout=10)
                if response.status_code == 200:
                    update_info = response.json()
                    # Schedule UI update on main thread
                    self.app.root.after(0, lambda: self.handle_update_response(update_info))
            except Exception as e:
                print(f"Update check failed: {e}")
                # No need to show error to user for failed update checks
        
        # Start update check in background thread
        update_thread = threading.Thread(target=update_check, daemon=True)
        update_thread.start()
    
    def handle_update_response(self, update_info):
        """Handle the update response on main thread"""
        latest_version = update_info.get('latest_version')
        download_url = update_info.get('download_url')
        release_notes = update_info.get('release_notes', '')
        mandatory = update_info.get('mandatory', False)
        
        if version.parse(latest_version) > version.parse(self.current_version):
            self.notify_update_available(latest_version, download_url, release_notes, mandatory)
    
    def notify_update_available(self, new_version, download_url, release_notes, mandatory):
        """Notify user about available update"""
        if mandatory:
            # Force update for mandatory updates
            message = f"MANDATORY UPDATE REQUIRED\n\nNew version {new_version} is available.\n\nThis update contains critical security fixes and is required to continue using the application.\n\nRelease Notes:\n{release_notes}\n\nDo you want to update now?"
            response = messagebox.askyesno("Mandatory Update Required", message)
            if response:
                self.download_and_install_update(download_url, new_version)
            else:
                messagebox.showwarning("Update Required", "You must update to continue using the application.")
                self.app.root.quit()
        else:
            # Optional update - show non-blocking message
            message = f"New version {new_version} is available!\n\nRelease Notes:\n{release_notes}\n\nWould you like to update now?"
            response = messagebox.askyesno("Update Available", message)
            if response:
                self.download_and_install_update(download_url, new_version)
    
    def download_and_install_update(self, download_url, new_version):
        """Download and install the update"""
        try:
            # Show download progress
            progress_window = self.create_progress_window("Downloading update...")
            
            def download_thread():
                try:
                    # Download the update
                    response = requests.get(download_url, stream=True)
                    total_size = int(response.headers.get('content-length', 0))
                    
                    temp_dir = tempfile.gettempdir()
                    installer_path = os.path.join(temp_dir, f"MERQ_Timesheet_Update_{new_version}.exe")
                    
                    with open(installer_path, 'wb') as f:
                        downloaded = 0
                        for chunk in response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                                downloaded += len(chunk)
                                if total_size > 0:
                                    progress = (downloaded / total_size) * 100
                                    # Update progress on main thread
                                    self.app.root.after(0, lambda: self.update_progress(progress_window, progress))
                    
                    # Close progress window and launch installer on main thread
                    self.app.root.after(0, lambda: self.finalize_update(progress_window, installer_path))
                    
                except Exception as e:
                    # Show error on main thread
                    self.app.root.after(0, lambda: self.show_download_error(progress_window, str(e)))
            
            # Start download in background thread
            download_thread_instance = threading.Thread(target=download_thread, daemon=True)
            download_thread_instance.start()
            
        except Exception as e:
            messagebox.showerror("Update Error", f"Failed to start download: {str(e)}")
    
    def finalize_update(self, progress_window, installer_path):
        """Finalize the update process"""
        progress_window.destroy()
        
        # Launch installer
        try:
            subprocess.Popen([installer_path, '/SILENT', '/NORESTART'])
            # Close current application
            self.app.root.quit()
        except Exception as e:
            messagebox.showerror("Update Error", f"Failed to launch installer: {str(e)}")
    
    def show_download_error(self, progress_window, error_message):
        """Show download error"""
        progress_window.destroy()
        messagebox.showerror("Update Error", f"Failed to download update: {error_message}")
    
    def create_progress_window(self, title):
        """Create progress window for download"""
        progress_window = tk.Toplevel(self.app.root)
        progress_window.title(title)
        progress_window.geometry("300x100")
        progress_window.transient(self.app.root)
        progress_window.grab_set()
        
        ttk.Label(progress_window, text="Downloading update...").pack(pady=10)
        progress_bar = ttk.Progressbar(progress_window, orient=tk.HORIZONTAL, length=250, mode='determinate')
        progress_bar.pack(pady=10)
        
        return progress_window
    
    def update_progress(self, progress_window, value):
        """Update progress bar"""
        for widget in progress_window.winfo_children():
            if isinstance(widget, ttk.Progressbar):
                widget['value'] = value
                progress_window.update()

class TimesheetApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MERQ Consultancy - Ethiopian Timesheet Generator V1.0")
        self.root.geometry("1400x900")
        self.root.state('zoomed')
        
        # Initialize database manager
        self.db_manager = DatabaseManager()
        
        # Initialize email manager
        self.email_manager = EmailManager()
        
        # User session
        self.user_session = None
        
        # Initialize update manager
        self.update_manager = UpdateManager(self)
        
        # Configure modern theme
        ModernTheme.configure_styles()
        
        # Set window icon
        self.set_window_icon()
        
        # Get current Ethiopian date
        self.current_ethiopian_date = EthiopianDateConverter.get_current_ethiopian_date()
        
        # Initialize variables
        self.employee_name = tk.StringVar(value="PLEASE ENTER YOUR FULL NAME")
        self.selected_year = tk.IntVar(value=self.current_ethiopian_date['year'])
        self.selected_month = tk.StringVar(value=self.current_ethiopian_date['month_name'])
        
        # Project data storage
        self.projects = []
        self.daily_hours = {}
        
        # Leave data storage
        self.leave_data = {
            "vacation": {'entries': {}, 'total_var': tk.DoubleVar(value=0.0)},
            "sick_leave": {'entries': {}, 'total_var': tk.DoubleVar(value=0.0)},
            "holiday": {'entries': {}, 'total_var': tk.DoubleVar(value=0.0)},
            "personal_leave": {'entries': {}, 'total_var': tk.DoubleVar(value=0.0)},
            "bereavement": {'entries': {}, 'total_var': tk.DoubleVar(value=0.0)},
            "other": {'entries': {}, 'total_var': tk.DoubleVar(value=0.0)}
        }
        
        # Live clock
        self.current_time_var = tk.StringVar()
        self.current_eth_date_var = tk.StringVar()
        
        # Application state
        self.name_entered = False
        
        # Show disclaimer first
        self.show_disclaimer()
        
        # Bind close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Start update check after UI is fully loaded
        self.root.after(5000, self.update_manager.check_for_updates_async)  # Check 5 seconds after startup
    
    def show_disclaimer(self):
        """Show disclaimer popup before main application"""
        disclaimer_window = tk.Toplevel(self.root)
        disclaimer_window.title("Important Disclaimer - አስፈላጊ ማስጠንቀቂያ")
        disclaimer_window.geometry("800x600")
        disclaimer_window.transient(self.root)
        disclaimer_window.grab_set()
        disclaimer_window.resizable(False, False)
        
        # Center the window
        disclaimer_window.update_idletasks()
        x = (disclaimer_window.winfo_screenwidth() // 2) - (800 // 2)
        y = (disclaimer_window.winfo_screenheight() // 2) - (600 // 2)
        disclaimer_window.geometry(f"800x600+{x}+{y}")
        
        # Main frame
        main_frame = ttk.Frame(disclaimer_window, style="Card.TFrame", padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(main_frame, text="⚠️ IMPORTANT DECLARATION & DISCLAIMER", 
                               style="Title.TLabel", foreground=ModernTheme.ACCENT)
        title_label.pack(pady=(0, 20))
        
        subtitle_label = ttk.Label(main_frame, text="አስፈላጊ መግለጫ እና ማስጠንቀቂያ", 
                                  style="Subtitle.TLabel")
        subtitle_label.pack(pady=(0, 30))
        
        # Disclaimer text
        disclaimer_text = """
በዚህ መተግበሪያ በሚያስገቡት ሁሉም መረጃዎች እና የስራ ሰዓት ማስታወሻዎች ላይ የሚከተለውን እውነት መሆኑን እገልጻለሁ፡

1. ይህ መተግበሪያ ለግል አጠቃቀሜ ብቻ ነው።
2. የምከፍለው ሁሉም መረጃ እና የስራ ሰዓት ማስታወሻዎች የራሴ ብቻ ናቸው።
3. ሁሉም መረጃዎች እውነት እና ትክክለኛ ናቸው።
4. ምንም ዓይነት የሌላ ሰው መረጃ አልገባም።
5. የምለው ሁሉ ከእውነታ በኋላ የተገኘ እና በትክክል የሰራሁበት ስራ ላይ የተመሰረተ ነው።

I hereby declare and confirm that:

1. This application is for my personal use only.
2. All information and timesheet entries I provide are my own.
3. All information is true and accurate.
4. I have not entered any information belonging to another person.
5. Everything I declare represents a reasonable after-the-fact estimate based on actual work performed by me.

By clicking "I AGREE & CONTINUE", you confirm that you understand and accept these terms.
        """
        
        text_widget = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD, font=("Arial", 11), 
                                              height=15, bg=ModernTheme.CARD_BG, relief="solid",
                                              borderwidth=1, padx=10, pady=10)
        text_widget.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        text_widget.insert(tk.END, disclaimer_text)
        text_widget.config(state=tk.DISABLED)
        
        # Agree button
        def on_agree():
            disclaimer_window.destroy()
            self.show_login_window()
        
        agree_button = ttk.Button(main_frame, text="✅ I AGREE & CONTINUE / እስማማለሁ እና ቀጥል", 
                                 command=on_agree, style="Success.TButton")
        agree_button.pack(pady=10)
        
        # Exit button
        exit_button = ttk.Button(main_frame, text="❌ EXIT APPLICATION / መተግበሪያውን ዝጋ", 
                                command=self.root.quit, style="Danger.TButton")
        exit_button.pack(pady=5)
    
    def show_login_window(self):
        """Show login window after disclaimer"""
        LoginWindow(self.root, self.db_manager, self.on_login_success)
    
    def on_login_success(self, user_data):
        """Handle successful login"""
        self.user_session = UserSession(user_data)
        self.create_widgets()
        self.add_default_project()
        self.update_calendar_display()
        self.update_clock()
        self.prevent_future_dates()
        
        # Update UI with user info
        self.update_user_display()
        
        # AUTO-UNLOCK: Set name_entered to True since user is logged in
        self.name_entered = True
        
        # Enable all input fields immediately
        self.enable_all_input_fields()
        
        
    def enable_all_input_fields(self):
        """Enable all input fields after successful login"""
        if hasattr(self, 'timesheet_frame'):
            # Enable project entries
            for project in self.projects:
                for day in project['entries']:
                    entry = project['entries'][day]
                    if hasattr(entry, 'tk'):
                        # Get the actual entry widget if it exists
                        widget = self.get_widget_by_variable(entry)
                        if widget:
                            widget.config(state=tk.NORMAL, background="white")
            
            # Enable leave entries
            for leave_key in self.leave_data:
                for day in self.leave_data[leave_key]['entries']:
                    entry = self.leave_data[leave_key]['entries'][day]
                    if hasattr(entry, 'tk'):
                        widget = self.get_widget_by_variable(entry)
                        if widget:
                            widget.config(state=tk.NORMAL, background="white")
        
        # Update any other locked UI elements
        self.update_calendar_display()

    def get_widget_by_variable(self, var):
        """Helper method to find widget by its variable"""
        # This is a simplified approach - you might need to track widgets differently
        for widget in self.timesheet_frame.winfo_children():
            if hasattr(widget, 'config'):
                try:
                    if hasattr(widget, 'cget') and widget.cget('textvariable') == str(var):
                        return widget
                except:
                    continue
        return None        
        
    
    def update_user_display(self):
        """Update UI with logged-in user information"""
        if not self.user_session:
            return
        
        # Update employee name automatically
        self.employee_name.set(self.user_session.full_name)
    
    def set_window_icon(self):
        """Set window icon from merq.png"""
        try:
            if os.path.exists("merq.png"):
                img = Image.open("merq.png")
                img = img.resize((32, 32), Image.Resampling.LANCZOS)
                self.logo_img = ImageTk.PhotoImage(img)
                self.root.iconphoto(True, self.logo_img)
        except Exception as e:
            print(f"Could not load icon: {e}")
    
    def prevent_future_dates(self):
        """Prevent selection of future months/years"""
        current_year = self.current_ethiopian_date['year']
        current_month = self.current_ethiopian_date['month']
        
        # Update combobox values to only allow current and past dates
        if hasattr(self, 'year_combo'):
            self.year_combo['values'] = [year for year in range(2010, current_year + 1)]
        
        # If current month is selected, only allow months up to current month
        if self.selected_year.get() == current_year and hasattr(self, 'month_combo'):
            available_months = EthiopianDateConverter.MONTHS_AMHARIC[:current_month]
            self.month_combo['values'] = available_months
    
    def create_widgets(self):
        # Configure root background
        self.root.configure(background=ModernTheme.BACKGROUND)
        
        # Main container
        main_container = ttk.Frame(self.root, style="Modern.TFrame")
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Header with logo and time
        header_frame = ttk.Frame(main_container, style="Card.TFrame", height=80)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        header_frame.pack_propagate(False)
        
        # Logo and title
        logo_title_frame = ttk.Frame(header_frame, style="Card.TFrame")
        logo_title_frame.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=10)
        
        # Logo
        try:
            if os.path.exists("merq.png"):
                img = Image.open("merq.png")
                img = img.resize((50, 50), Image.Resampling.LANCZOS)
                self.header_logo_img = ImageTk.PhotoImage(img)
                logo_label = ttk.Label(logo_title_frame, image=self.header_logo_img, 
                                      background=ModernTheme.CARD_BG)
                logo_label.pack(side=tk.LEFT, padx=(0, 15))
        except:
            pass
        
        title_frame = ttk.Frame(logo_title_frame, style="Card.TFrame")
        title_frame.pack(side=tk.LEFT, fill=tk.Y)
        
        ttk.Label(title_frame, text="MERQ CONSULTANCY", style="Title.TLabel").pack(anchor=tk.W)
        ttk.Label(title_frame, text="ወርሃዊ የስራ ሰዓት መከታተያ / Monthly Timesheet Tracker ", 
                 style="Subtitle.TLabel").pack(anchor=tk.W)
        
        # Current time and date display
        time_frame = ttk.Frame(header_frame, style="Card.TFrame")
        time_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=20, pady=10)
        
        current_eth_date = f"{self.current_ethiopian_date['day']} {self.current_ethiopian_date['month_name']} {self.current_ethiopian_date['year']}"
        ttk.Label(time_frame, text=current_eth_date, style="Subtitle.TLabel", 
                 foreground=ModernTheme.SECONDARY, font=("Arial", 11, "bold")).pack(anchor=tk.E)
        ttk.Label(time_frame, textvariable=self.current_time_var, style="Subtitle.TLabel",
                 foreground=ModernTheme.ACCENT, font=("Arial", 10)).pack(anchor=tk.E)
        ttk.Label(time_frame, textvariable=self.current_eth_date_var, style="Subtitle.TLabel",
                 foreground=ModernTheme.SUCCESS, font=("Arial", 9)).pack(anchor=tk.E)
        
        # Main content area with paned window (1:4 ratio - left:right)
        content_paned = ttk.PanedWindow(main_container, orient=tk.HORIZONTAL)
        content_paned.pack(fill=tk.BOTH, expand=True)
        
        # Left sidebar (1/5 width)
        left_sidebar = ttk.Frame(content_paned, style="Card.TFrame", width=280)
        content_paned.add(left_sidebar, weight=1)
        
        # Right content area (4/5 width)
        right_content = ttk.Frame(content_paned, style="Modern.TFrame")
        content_paned.add(right_content, weight=4)
        
        # Build left sidebar
        self.build_left_sidebar(left_sidebar)
        
        # Build right content area
        self.build_right_content(right_content)
        
        # Footer
        self.create_footer(main_container)
    
    def create_footer(self, parent):
        """Create footer with developer information"""
        footer_frame = ttk.Frame(parent, style="Card.TFrame", height=30)
        footer_frame.pack(fill=tk.X, pady=(10, 0))
        footer_frame.pack_propagate(False)
        
        # Copyright text
        copyright_text = "© 2025 MERQ Consultancy PLC - Developed by Information Systems & Digital Health Unit (ISDHU) - Version 1.0.0.0"
        copyright_label = ttk.Label(footer_frame, text=copyright_text, 
                                   style="Dark.TLabel", font=("Arial", 8),
                                   foreground=ModernTheme.TEXT_MUTED)
        copyright_label.pack(side=tk.LEFT, padx=20, pady=5)
        
        # Website link
        def open_website():
            webbrowser.open("https://app.merqconsultancy.org/")
        
        website_label = ttk.Label(footer_frame, text="🌐 Visit Cloud App", 
                                 style="Dark.TLabel", font=("Arial", 8, "underline"),
                                 foreground=ModernTheme.SECONDARY, cursor="hand2")
        website_label.pack(side=tk.RIGHT, padx=20, pady=5)
        website_label.bind("<Button-1>", lambda e: open_website())
    
    def build_left_sidebar(self, parent):
        """Build the left sidebar with controls (1/5 of screen)"""
        # Create scrollable sidebar
        sidebar_container = ttk.Frame(parent, style="Card.TFrame")
        sidebar_container.pack(fill=tk.BOTH, expand=True)
        
        # Vertical scrollbar for sidebar
        v_scrollbar = ttk.Scrollbar(sidebar_container, style="Modern.Vertical.TScrollbar")
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Canvas for scrollable content
        sidebar_canvas = tk.Canvas(sidebar_container, bg=ModernTheme.CARD_BG,
                                yscrollcommand=v_scrollbar.set, highlightthickness=0)
        sidebar_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.config(command=sidebar_canvas.yview)
        
        # Scrollable frame
        scrollable_frame = ttk.Frame(sidebar_canvas, style="Card.TFrame")
        sidebar_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        def configure_scroll_region(event):
            sidebar_canvas.configure(scrollregion=sidebar_canvas.bbox("all"))
        
        scrollable_frame.bind("<Configure>", configure_scroll_region)
        
        # Enhanced Employee info section with user data
        info_frame = ttk.LabelFrame(scrollable_frame, text="👤 ሰራተኛ/አማካሪ መረጃ / Employee Information", 
                                style="Card.TFrame", padding=15)
        info_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Display user information
        if self.user_session:
            user_info = [
                f"ስም / Name: {self.user_session.full_name}",
                f"መጠይቅ / Position: {self.user_session.position}",
                f"የስራ ክፍል / Department: {self.user_session.department}",
                f"የበላይ ሹም / Supervisor: {self.user_session.supervisor_name}",
                f"የበላይ ሹም መጠይቅ / Supervisor Position: {self.user_session.supervisor_position_title}",
                f"የሰራተኛ መለያ / Employee ID: {self.user_session.employee_id}"
            ]
            
            for info in user_info:
                label = ttk.Label(info_frame, text=info, style="Dark.TLabel", 
                                font=("Arial", 9))
                label.pack(anchor=tk.W, pady=2)
        
        # Year selection frame
        year_frame = ttk.Frame(info_frame, style="Card.TFrame")
        year_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(year_frame, text="ዓመት / Year:", style="Dark.TLabel",
                font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        
        self.year_combo = ttk.Combobox(year_frame, textvariable=self.selected_year, 
                                values=list(range(2010, 2031)), width=12, 
                                font=("Arial", 10), style="Modern.TCombobox")
        self.year_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.year_combo.bind('<<ComboboxSelected>>', self.on_date_selection)
        
        # Month selection frame
        month_frame = ttk.Frame(info_frame, style="Card.TFrame")
        month_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(month_frame, text="ወር / Month:", style="Dark.TLabel",
                font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        
        self.month_combo = ttk.Combobox(month_frame, textvariable=self.selected_month, 
                                values=EthiopianDateConverter.MONTHS_AMHARIC, width=12, 
                                font=("Arial", 10), style="Modern.TCombobox")
        self.month_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.month_combo.bind('<<ComboboxSelected>>', self.on_date_selection)
        
        # Prefill Hours Button
        prefill_frame = ttk.Frame(info_frame, style="Card.TFrame")
        prefill_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(prefill_frame, text="🕐 ነባሪ ሰዓቶችን አስገባ / Prefill Default Hours", 
                command=self.prefill_default_hours, style="Primary.TButton").pack(fill=tk.X)
        
        # Interactive calendar
        calendar_frame = ttk.LabelFrame(scrollable_frame, text="🗓 የኢትዮጵያ የቀን መቁጠሪያ / Ethiopian Calendar", 
                                    style="Card.TFrame", padding=10)
        calendar_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Simple calendar display
        calendar_display_frame = ttk.Frame(calendar_frame, style="Card.TFrame")
        calendar_display_frame.pack(fill=tk.X)
        
        # Month and year display
        month_year_label = ttk.Label(calendar_display_frame, 
                                text=f"{self.selected_month.get()} {self.selected_year.get()}",
                                style="Title.TLabel", font=("Arial", 12, "bold"))
        month_year_label.pack(pady=5)
        
        # Projects management
        projects_frame = ttk.LabelFrame(scrollable_frame, text="📄ፕሮጀክቶች አስተዳደር / Projects Management", 
                                    style="Card.TFrame", padding=15)
        projects_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.projects_container = ttk.Frame(projects_frame, style="Card.TFrame")
        self.projects_container.pack(fill=tk.X)
        
        ttk.Button(projects_frame, text="➕ ተጨማሪ ፕሮጀክት ጨምር / Add Project", 
                command=self.add_project, style="Primary.TButton").pack(fill=tk.X, pady=5)
        
        # Action buttons
        action_frame = ttk.LabelFrame(scrollable_frame, text="📝 ድርጊቶች / Actions", style="Card.TFrame", padding=15)
        action_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Button(action_frame, text="🔎 ማጠቃለያ ሰሌዳ አሳይ / Preview Timesheet Summary", 
                command=self.preview_timesheet, style="Primary.TButton").pack(fill=tk.X, pady=5)
        
        ttk.Button(action_frame, text="✅ታይምሺት ኤከሴል አውርድ 🔽 Download Timesheet Excel", 
                command=self.export_to_excel, style="Success.TButton").pack(fill=tk.X, pady=5)
        
        ttk.Button(action_frame, text="📧 ታይምሺት ለኤችአር አስቀምጥ / Submit Timesheet to HR", 
                command=self.submit_timesheet, style="Success.TButton").pack(fill=tk.X, pady=5)
        
        ttk.Button(action_frame, text="❌ አጽዳ / Clear All", 
                command=self.clear_all, style="Danger.TButton").pack(fill=tk.X, pady=5)
        
        # Help section
        help_frame = ttk.LabelFrame(scrollable_frame, text="❓እገዛ / Help", style="Card.TFrame", padding=15)
        help_frame.pack(fill=tk.X, pady=(0, 15))
        
        help_text = """የሰዓት ሰሌዳ አጠቃቀም:
    1. ስምዎን ያስገቡ እና ዓመት/ወር ይምረጡ
    2. 'ፕሮጀክት ጨምር' በማለት ፕሮጀክቶችን ያክሉ
    3. ለእያንዳንድ ፕሮጀክት የቀን ሰዓቶችን ያስገቡ
    4. ከፈለጉ የፈቃድ ሰዓቶችን ያስገቡ
    5. ለመፈተሽ 'ማጠቃለያ ሰሌዳ አሳይ' ይጫኑ
    6. ለማስቀመጥ/ለማውረድ 'ታይምሺት ኤከሴል አውርድ'ን ይጫኑ  
    7. ካወረዱ በኋላ የወረደውን ወደ haymanot.a@merqconsultancy.org ይላኩ።  

    ይህ ስርዓት የተገነባው በመርክ ኮንሱልታንሲ በኢንፎርሜሽን ሲስተምስ እና ዲጂታል የጤና ክፍል ነው።
    ማንኛውም ችግር ካጋጠመዎት እባክዎ support@merqconsultancy.org ን ያነጋግሩ።    


    How to use this timesheet:
    1. Enter your name and select year/month
    2. Add projects using 'Add Project' button
    3. Enter daily hours for each project
    4. Enter leave hours if applicable
    5. Click 'Preview Timesheet Summary' to review
    6. Click 'Download TImesheet Excel' to save
    7. After you've download it send the downloaded to haymanot.a@merqconsultancy.org

    This system is developed by MERQ Consultancy's Information systems & Digital Health Unit.
    If you have any issues please contact support@merqconsultancy.org


    """
        
        help_text_widget = scrolledtext.ScrolledText(help_frame, wrap=tk.WORD, font=("Arial", 9), 
                                                height=8, bg=ModernTheme.CARD_BG, relief="flat",
                                                foreground=ModernTheme.TEXT_DARK)
        help_text_widget.pack(fill=tk.BOTH, expand=True)
        help_text_widget.insert(tk.END, help_text)
        help_text_widget.config(state=tk.DISABLED)
        
        # Update the canvas scrollregion after all widgets are added
        def update_scrollregion():
            sidebar_canvas.configure(scrollregion=sidebar_canvas.bbox("all"))
        
        scrollable_frame.after(100, update_scrollregion)

    def on_name_change(self, *args):
        """Handle name change to enable/disable input fields"""
        name = self.employee_name.get().strip()
        self.name_entered = name != "" and name != "PLEASE ENTER YOUR FULL NAME"
        
        # Update calendar display to enable/disable fields
        if hasattr(self, 'timesheet_frame'):
            self.update_calendar_display()

#    def on_date_selection(self, event=None):
#        """Handle date selection changes"""
#        self.update_calendar_display()
#        self.prevent_future_dates()


    def on_date_selection(self, event=None):
        """Handle date selection changes"""
        self.update_calendar_display()
        self.prevent_future_dates()
        
        # Update default project allocated hours when month/year changes
        if self.projects and self.projects[0]['name_var'].get() == "MERQ Internal":
            dynamic_hours = self.calculate_total_working_hours()
            self.projects[0]['hours_var'].set(dynamic_hours)

    def build_right_content(self, parent):
        """Build the right content area with timesheet (4/5 of screen)"""
        # Timesheet container with scrollbars
        timesheet_container = ttk.Frame(parent, style="Modern.TFrame")
        timesheet_container.pack(fill=tk.BOTH, expand=True)
        
        # Create horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(timesheet_container, orient=tk.HORIZONTAL, 
                                style="Modern.Horizontal.TScrollbar")
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Create vertical scrollbar
        v_scrollbar = ttk.Scrollbar(timesheet_container, style="Modern.Vertical.TScrollbar")
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Create canvas for scrolling
        self.canvas = tk.Canvas(timesheet_container, 
                            xscrollcommand=h_scrollbar.set,
                            yscrollcommand=v_scrollbar.set,
                            bg=ModernTheme.CONTENT_BG, highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Configure scrollbars
        h_scrollbar.config(command=self.canvas.xview)
        v_scrollbar.config(command=self.canvas.yview)
        
        # Create frame inside canvas
        self.timesheet_frame = ttk.Frame(self.canvas, style="Modern.TFrame")
        self.canvas_window = self.canvas.create_window((0, 0), window=self.timesheet_frame, anchor="nw")
        
        # Bind events for scrolling and resizing
        self.timesheet_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)

    def on_frame_configure(self, event):
        """Update scrollregion when frame size changes"""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_canvas_configure(self, event):
        """Update canvas window width when canvas is resized"""
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def update_clock(self):
        """Update the current time display"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.current_time_var.set(f"Gregorian: {current_time}")
        
        # Update Ethiopian date
        eth_date = EthiopianDateConverter.get_current_ethiopian_date()
        eth_display = f"ዛሬ ቀኑ: {eth_date['month_name']} {eth_date['day']} {eth_date['year']}"
        self.current_eth_date_var.set(eth_display)
        
        self.root.after(1000, self.update_clock)

#    def add_default_project(self):
#        """Add MERQ Internal project by default"""
#        self.add_project("MERQ Internal", 192.0)

    def add_default_project(self):
        """Add MERQ Internal project with dynamic allocated hours"""
        # Calculate dynamic hours based on selected month
        dynamic_hours = self.calculate_total_working_hours()
        self.add_project("MERQ Internal", dynamic_hours)        

    def add_project(self, name="", hours=0.0):
        """Add a new project entry with total display - IMPROVED"""
        project_frame = ttk.Frame(self.projects_container, style="Card.TFrame")
        project_frame.pack(fill=tk.X, pady=2)
        
        project_index = len(self.projects)
        
        # Project name
        project_name = tk.StringVar(value=name)
        project_name.trace('w', lambda *args: self.update_project_total_displays())
        ttk.Label(project_frame, text="ፕሮጀክት ስም:", style="Dark.TLabel", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        project_entry = ttk.Entry(project_frame, textvariable=project_name, width=15, style="Modern.TEntry")
        project_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # Hours allocated
        hours_allocated = tk.DoubleVar(value=hours)
        hours_allocated.trace('w', lambda *args: self.update_project_total_displays())  # Add trace for allocated hours
        ttk.Label(project_frame, text="⏰ የተመደበ:", style="Dark.TLabel", font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 5))
        hours_entry = ttk.Entry(project_frame, textvariable=hours_allocated, width=6, style="Modern.TEntry")
        hours_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        # Total hours worked (LIVE UPDATING)
        total_var = tk.DoubleVar(value=0.0)
        total_label = ttk.Label(project_frame, textvariable=total_var, style="Dark.TLabel", 
                            foreground=ModernTheme.SUCCESS, font=("Arial", 9, "bold"))
        total_label.pack(side=tk.LEFT, padx=(0, 5))
        ttk.Label(project_frame, text="ሰአት የሰራሁ", style="Dark.TLabel", font=("Arial", 9)).pack(side=tk.LEFT, padx=(0, 10))
        
        # Progress indicator (optional - shows percentage of allocated hours used)
        progress_var = tk.StringVar(value="0%")
        progress_label = ttk.Label(project_frame, textvariable=progress_var, style="Dark.TLabel", 
                                foreground=ModernTheme.INFO, font=("Arial", 8))
        progress_label.pack(side=tk.LEFT, padx=(0, 10))
        
        # Remove button
        remove_btn = ttk.Button(project_frame, text="🗑", 
                            command=lambda: self.remove_project(project_index, project_frame),
                            style="Danger.TButton", width=3)
        remove_btn.pack(side=tk.LEFT)
        
        project_data = {
            'frame': project_frame,
            'name_var': project_name,
            'hours_var': hours_allocated,
            'entries': {},
            'total_var': total_var,
            'total_label': total_label,
            'progress_var': progress_var,
            'progress_label': progress_label
        }
        
        self.projects.append(project_data)
        self.daily_hours[project_index] = {}
        
        # Force update of calendar display to show new project
        self.update_calendar_display()
        
        # Update trace for allocated hours to also update progress
        hours_allocated.trace('w', lambda *args: self.update_project_total_displays())

    def remove_project(self, project_index, frame):
        """Remove a project"""
        if 0 <= project_index < len(self.projects):
            if len(self.projects) <= 1:
                messagebox.showwarning("Warning", "At least one project must remain")
                return
                
            # Remove from data structures
            if project_index in self.daily_hours:
                del self.daily_hours[project_index]
            
            # Update indices for remaining projects
            new_daily_hours = {}
            for i, (idx, data) in enumerate(self.daily_hours.items()):
                if idx > project_index:
                    new_daily_hours[i-1] = data
                elif idx < project_index:
                    new_daily_hours[idx] = data
            self.daily_hours = new_daily_hours
            
            # Remove project from list
            self.projects.pop(project_index)
            
            # Destroy frame
            frame.destroy()
            
            # Update display
            self.update_calendar_display()

    def safe_float_convert(self, value):
        """Safely convert string to float, return 0.0 if invalid"""
        try:
            if value == "" or value is None:
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def validate_hours_input(self, P, day_total=0.0):
        """Validate hours input to ensure total doesn't exceed 24 hours"""
        if P == "" or P == ".":
            return True
        try:
            value = float(P)
            # Check if total for the day would exceed 24 hours
            if day_total + value > 24:
                return False
            return value >= 0
        except ValueError:
            return False

    def prefill_default_hours(self):
        """Prefill default hours based on weekday"""
        # REMOVE this check:
        # if not self.name_entered:
        #     messagebox.showerror("Error", "Please enter your full name first before prefilling hours.")
        #     return
        
        year = self.selected_year.get()
        month_name = self.selected_month.get()
        month_index = EthiopianDateConverter.MONTHS_AMHARIC.index(month_name) + 1
        month_days = EthiopianDateConverter.get_ethiopian_month_days(year, month_index)
        
        for day in range(1, month_days + 1):
            weekday_index = EthiopianDateConverter.get_ethiopian_weekday(year, month_index, day)
            
            # Set default hours based on weekday
            if weekday_index < 4:  # Monday-Thursday: 8 hours
                default_hours = 8.0
            elif weekday_index == 4:  # Friday: 8 hours
                default_hours = 8.0
            elif weekday_index == 5:  # Saturday: 4 hours
                default_hours = 4.0
            else:  # Sunday: 0 hours
                default_hours = 0.0
            
            # Fill the first project with default hours
            if self.projects and day in self.projects[0]['entries']:
                self.projects[0]['entries'][day].set(str(default_hours))
        
        messagebox.showinfo("Success", "Default hours have been prefilled based on weekdays:\n\n"
                                    "ሰኞ-ሐሙስ / Mon-Thu: 8 hours\n"
                                    "ዓርብ / Friday: 8 hours\n"
                                    "ቅዳሜ / Saturday: 4 hours\n"
                                    "እሁድ / Sunday: 0 hours")

    def update_calendar_display(self):
        """Update the calendar display based on selected year and month"""
        # Clear existing calendar
        for widget in self.timesheet_frame.winfo_children():
            widget.destroy()
        
        year = self.selected_year.get()
        month_name = self.selected_month.get()
        month_index = EthiopianDateConverter.MONTHS_AMHARIC.index(month_name) + 1
        
        # Get month days
        month_days = EthiopianDateConverter.get_ethiopian_month_days(year, month_index)
        
        # Create main title
        title_label = ttk.Label(self.timesheet_frame, 
                            text="የመርክ ኮንሱልታንሲ ሰራተኛ/አማካሪ የስራ ሪከርድ (በቀን ሰዓት) / MERQ Consultancy's Employee/Consultant Work Record (Hours per day)",
                            font=("Arial", 12, "bold"), background=ModernTheme.CONTENT_BG,
                            foreground=ModernTheme.TEXT_DARK)
        title_label.grid(row=0, column=0, columnspan=month_days+4, pady=(0, 10), sticky=tk.W+tk.E)
        
        # Create header rows
        headers_row1 = ["", "", "የቀን ቁጥር / Day Number"] + [""] * (month_days - 0)
        headers_row2 = ["የሰራተኛ/የአማካሪ ስም", "ፕሮጀክት ስም/መለያ", "ቀን/Date"] + [f"{i+1}" for i in range(0, month_days)]
        
        for col, header in enumerate(headers_row1):
            label = ttk.Label(self.timesheet_frame, text=header, borderwidth=1, relief="solid", 
                            width=12, background=ModernTheme.SECONDARY, foreground=ModernTheme.TEXT_LIGHT, 
                            font=("Arial", 9, "bold"))
            label.grid(row=1, column=col, sticky=tk.W+tk.E)
        
        for col, header in enumerate(headers_row2):
            label = ttk.Label(self.timesheet_frame, text=header, borderwidth=1, relief="solid", 
                            width=12, background=ModernTheme.SECONDARY, foreground=ModernTheme.TEXT_LIGHT, 
                            font=("Arial", 9, "bold"))
            label.grid(row=2, column=col, sticky=tk.W+tk.E)
        
        # Employee name row
        ttk.Label(self.timesheet_frame, text="Employee/Consultant Name:", borderwidth=1, relief="solid", 
                width=35, background=ModernTheme.LIGHT, foreground=ModernTheme.TEXT_DARK).grid(row=3, column=0, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, textvariable=self.employee_name, borderwidth=1, relief="solid", 
                width=60, background="white", foreground=ModernTheme.TEXT_DARK).grid(row=3, column=1, columnspan=35, sticky=tk.W+tk.E)
        
        # Date and weekday rows
        date_row = 4
        weekday_row = 5
        
        # Empty cells for project area
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background=ModernTheme.LIGHT).grid(row=date_row, column=0, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background=ModernTheme.LIGHT).grid(row=date_row, column=1, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="ቀን/Date", borderwidth=1, relief="solid", 
                width=40, background=ModernTheme.LIGHT, foreground=ModernTheme.TEXT_DARK).grid(row=date_row, column=0, columnspan=3, sticky=tk.W+tk.E)
        
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background=ModernTheme.LIGHT).grid(row=weekday_row, column=0, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background=ModernTheme.LIGHT).grid(row=weekday_row, column=1, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="ዕለት/Day", borderwidth=1, relief="solid", 
                width=40, background=ModernTheme.LIGHT, foreground=ModernTheme.TEXT_DARK).grid(row=weekday_row, column=0, columnspan=3, sticky=tk.W+tk.E)
        
        # Fill dates and weekdays
        for day in range(1, month_days + 1):
            eth_date = EthiopianDateConverter.format_ethiopian_date(year, month_index, day)
            weekday_index = EthiopianDateConverter.get_ethiopian_weekday(year, month_index, day)
            weekday_name = EthiopianDateConverter.WEEK_DAYS_AMHARIC[weekday_index]
            
            # Color weekend days light green
            bg_color = "#D4EDDA" if weekday_index >= 5 else "white"
            fg_color = ModernTheme.TEXT_DARK
            
            ttk.Label(self.timesheet_frame, text=eth_date, borderwidth=1, relief="solid", 
                    width=12, background=bg_color, font=("Arial", 8), foreground=fg_color).grid(row=date_row, column=day+2, sticky=tk.W+tk.E)
            ttk.Label(self.timesheet_frame, text=weekday_name, borderwidth=1, relief="solid", 
                    width=12, background=bg_color, font=("Arial", 9, "bold"), foreground=fg_color).grid(row=weekday_row, column=day+2, sticky=tk.W+tk.E)
        
        # Project rows starting from row 6
        start_row = 6
        for project_idx, project in enumerate(self.projects):
            row = start_row + project_idx
            
            # Project name
            ttk.Label(self.timesheet_frame, text=project['name_var'].get(), borderwidth=1, relief="solid", 
                    width=15, background="#FFF3CD", foreground=ModernTheme.TEXT_DARK).grid(row=row, column=0, sticky=tk.W+tk.E)
            
            # Direct label
            ttk.Label(self.timesheet_frame, text="ቀጥታ የሰራሁባቸው ሰዓቶች 👉🏾", borderwidth=1, relief="solid", 
                    width=40, background="#FFF3CD", foreground=ModernTheme.TEXT_DARK).grid(row=row, column=1, sticky=tk.W+tk.E)
            
            # Empty cell for date label
            ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                    width=12, background="#FFF3CD").grid(row=row, column=2, sticky=tk.W+tk.E)
            
            # Create entry for each day
            for day in range(1, month_days + 1):
                weekday_index = EthiopianDateConverter.get_ethiopian_weekday(year, month_index, day)
                bg_color = "#D4EDDA" if weekday_index >= 5 else "white"
                
                hours_var = tk.StringVar(value="")
                entry = ttk.Entry(self.timesheet_frame, textvariable=hours_var, width=8,
                                justify=tk.CENTER, background=bg_color, font=("Arial", 9),
                                foreground=ModernTheme.TEXT_DARK)
                entry.grid(row=row, column=day+2, sticky=tk.W+tk.E)
                
                # Enable/disable based on name entry
                if not self.name_entered:
                    entry.config(state=tk.DISABLED, background="#f0f0f0")
                else:
                    entry.config(state=tk.NORMAL, background=bg_color)
                
                # Validate input to allow only numbers and check total hours
                def validate_input(P, current_day=day, current_project=project_idx):
                    if not self.name_entered:
                        return False
                    
                    if P == "" or P == ".":
                        return True
                    try:
                        value = float(P)
                        if value < 0:
                            return False
                        
                        # Calculate current total for the day across all projects
                        day_total = 0.0
                        for proj in self.projects:
                            day_var = proj['entries'].get(current_day)
                            if day_var:
                                day_value = day_var.get()
                                if day_value and day_value != "":
                                    day_total += self.safe_float_convert(day_value)
                        
                        # If this entry is being updated, subtract its current value and add new value
                        current_value = self.safe_float_convert(hours_var.get())
                        day_total = day_total - current_value + value
                        
                        # Check if total exceeds 24 hours
                        if day_total > 24:
                            messagebox.showwarning("Warning", 
                                                f"Total hours for day {current_day} cannot exceed 24 hours!\n"
                                                f"Current total: {day_total:.1f} hours")
                            return False
                        return True
                    except ValueError:
                        return False
                
                vcmd = (self.root.register(validate_input), '%P')
                entry.config(validate="key", validatecommand=vcmd)
                
                # Store reference
                project['entries'][day] = hours_var
                self.daily_hours[project_idx][day] = hours_var
                
                # Bind to update totals with safe conversion
                #hours_var.trace('w', lambda *args, p_idx=project_idx: self.update_project_totals(p_idx))
                hours_var.trace('w', lambda *args, p_idx=project_idx: self.update_project_total_displays())
        
        # Totals row for direct work
        total_row = start_row + len(self.projects)
        ttk.Label(self.timesheet_frame, text="TOTAL ቀጥታ", borderwidth=1, relief="solid", 
                width=15, background=ModernTheme.INFO, foreground=ModernTheme.TEXT_LIGHT, 
                font=("Arial", 9, "bold")).grid(row=total_row, column=0, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background=ModernTheme.INFO).grid(row=total_row, column=1, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background=ModernTheme.INFO).grid(row=total_row, column=2, sticky=tk.W+tk.E)
        
        # Total cells for each day
        self.daily_total_vars = {}
        for day in range(1, month_days + 1):
            total_var = tk.DoubleVar(value=0.0)
            ttk.Label(self.timesheet_frame, textvariable=total_var, borderwidth=1, relief="solid", 
                    width=18, background=ModernTheme.INFO, foreground=ModernTheme.TEXT_LIGHT, 
                    font=("Arial", 9, "bold")).grid(row=total_row, column=day+2, sticky=tk.W+tk.E)
            self.daily_total_vars[day] = total_var
        
        # Leave section
        leave_start_row = total_row + 1
        leave_types = [
            ("የእረፍት ጊዜ / VACATION", "vacation"),
            ("የጤና እረፍት / SICK LEAVE", "sick_leave"),
            ("በዓል / HOLIDAY", "holiday"),
            ("የግል ፈቃድ / PERSONAL LEAVE", "personal_leave"),
            ("የሐዘን እረፍት / BEREAVEMENT", "bereavement"),
            ("ሌሎች / Other", "other")
        ]
        
        for i, (leave_name, leave_key) in enumerate(leave_types):
            row = leave_start_row + i
            
            ttk.Label(self.timesheet_frame, text="ፈቃድ *", borderwidth=1, relief="solid", 
                    width=15, background="#F8D7DA", foreground=ModernTheme.TEXT_DARK).grid(row=row, column=0, sticky=tk.W+tk.E)
            ttk.Label(self.timesheet_frame, text=leave_name, borderwidth=1, relief="solid", 
                    width=12, background="#F8D7DA", foreground=ModernTheme.TEXT_DARK).grid(row=row, column=1, sticky=tk.W+tk.E)
            ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                    width=12, background="#F8D7DA").grid(row=row, column=2, sticky=tk.W+tk.E)
            
            for day in range(1, month_days + 1):
                weekday_index = EthiopianDateConverter.get_ethiopian_weekday(year, month_index, day)
                bg_color = "#D4EDDA" if weekday_index >= 5 else "white"
                
                hours_var = tk.StringVar(value="")
                entry = ttk.Entry(self.timesheet_frame, textvariable=hours_var, width=8,
                                justify=tk.CENTER, background=bg_color, font=("Arial", 9),
                                foreground=ModernTheme.TEXT_DARK)
                entry.grid(row=row, column=day+2, sticky=tk.W+tk.E)
                
                # Enable/disable based on name entry
                if not self.name_entered:
                    entry.config(state=tk.DISABLED, background="#f0f0f0")
                else:
                    entry.config(state=tk.NORMAL, background=bg_color)
                
                # Validate input
                def validate_input(P, current_day=day):
                    if not self.name_entered:
                        return False
                    
                    if P == "" or P == ".":
                        return True
                    try:
                        value = float(P)
                        if value < 0:
                            return False
                        
                        # Calculate current total for the day across all projects and leave
                        day_total = 0.0
                        
                        # Add project hours
                        for proj in self.projects:
                            day_var = proj['entries'].get(current_day)
                            if day_var:
                                day_value = day_var.get()
                                if day_value and day_value != "":
                                    day_total += self.safe_float_convert(day_value)
                        
                        # Add leave hours
                        for lk in self.leave_data:
                            lv_var = self.leave_data[lk]['entries'].get(current_day)
                            if lv_var:
                                lv_value = lv_var.get()
                                if lv_value and lv_value != "":
                                    # For the current leave type being edited, use the new value
                                    if lk == leave_key:
                                        day_total += value
                                    else:
                                        day_total += self.safe_float_convert(lv_value)
                        
                        # Check if total exceeds 24 hours
                        if day_total > 24:
                            messagebox.showwarning("Warning", 
                                                f"Total hours for day {current_day} cannot exceed 24 hours!\n"
                                                f"Current total: {day_total:.1f} hours")
                            return False
                        return True
                    except ValueError:
                        return False
                
                vcmd = (self.root.register(validate_input), '%P')
                entry.config(validate="key", validatecommand=vcmd)
                
                self.leave_data[leave_key]['entries'][day] = hours_var
                hours_var.trace('w', lambda *args, lk=leave_key: self.update_leave_total(lk))
        
        # Total leave row
        total_leave_row = leave_start_row + len(leave_types)
        ttk.Label(self.timesheet_frame, text="ጠቅላላ ፈቃድ", borderwidth=1, relief="solid", 
                width=15, background="#F8D7DA", foreground=ModernTheme.TEXT_DARK, 
                font=("Arial", 9, "bold")).grid(row=total_leave_row, column=0, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background="#F8D7DA").grid(row=total_leave_row, column=1, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background="#F8D7DA").grid(row=total_leave_row, column=2, sticky=tk.W+tk.E)
        
        self.leave_total_vars = {}
        for day in range(1, month_days + 1):
            total_var = tk.DoubleVar(value=0.0)
            ttk.Label(self.timesheet_frame, textvariable=total_var, borderwidth=1, relief="solid", 
                    width=12, background="#F8D7DA", foreground=ModernTheme.TEXT_DARK, 
                    font=("Arial", 9, "bold")).grid(row=total_leave_row, column=day+2, sticky=tk.W+tk.E)
            self.leave_total_vars[day] = total_var
        
        # Grand total row
        grand_total_row = total_leave_row + 1
        ttk.Label(self.timesheet_frame, text="ጠቅላላ", borderwidth=1, relief="solid", 
                width=15, background="#FFEAA7", foreground=ModernTheme.TEXT_DARK, 
                font=("Arial", 10, "bold")).grid(row=grand_total_row, column=0, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background="#FFEAA7").grid(row=grand_total_row, column=1, sticky=tk.W+tk.E)
        ttk.Label(self.timesheet_frame, text="", borderwidth=1, relief="solid", 
                width=12, background="#FFEAA7").grid(row=grand_total_row, column=2, sticky=tk.W+tk.E)
        
        self.grand_total_vars = {}
        for day in range(1, month_days + 1):
            total_var = tk.DoubleVar(value=0.0)
            ttk.Label(self.timesheet_frame, textvariable=total_var, borderwidth=1, relief="solid", 
                    width=12, background="#FFEAA7", foreground=ModernTheme.TEXT_DARK, 
                    font=("Arial", 10, "bold")).grid(row=grand_total_row, column=day+2, sticky=tk.W+tk.E)
            self.grand_total_vars[day] = total_var
        
        # Summary section
        summary_start = grand_total_row + 2
        
        # Configure grid weights for responsiveness
        for col in range(month_days + 3):
            self.timesheet_frame.columnconfigure(col, weight=1)
        
        # Update all totals initially
        self.update_all_totals()

    def update_project_totals(self, project_idx):
        """Update totals when project hours change"""
        self.update_all_totals()

    def update_leave_total(self, leave_key):
        """Update total for a leave type"""
        self.update_all_totals()

    def update_all_totals(self):
        """Update all totals dynamically with safe conversion - enhanced with real-time updates"""
        year = self.selected_year.get()
        month_name = self.selected_month.get()
        month_index = EthiopianDateConverter.MONTHS_AMHARIC.index(month_name) + 1
        month_days = EthiopianDateConverter.get_ethiopian_month_days(year, month_index)
        
        # Update daily totals for direct work with safe conversion
        for day in range(1, month_days + 1):
            day_total = 0.0
            for project_idx in range(len(self.projects)):
                hours_var = self.projects[project_idx]['entries'].get(day)
                if hours_var:
                    hours_str = hours_var.get()
                    day_total += self.safe_float_convert(hours_str)
            
            if hasattr(self, 'daily_total_vars') and day in self.daily_total_vars:
                self.daily_total_vars[day].set(round(day_total, 2))
        
        # Update leave totals - FIXED: Ensure this always runs
        for day in range(1, month_days + 1):
            leave_total = 0.0
            for leave_key in self.leave_data:
                hours_var = self.leave_data[leave_key]['entries'].get(day)
                if hours_var:
                    hours_str = hours_var.get()
                    leave_total += self.safe_float_convert(hours_str)
            if hasattr(self, 'leave_total_vars') and day in self.leave_total_vars:
                self.leave_total_vars[day].set(round(leave_total, 2))
        
        # Update grand totals - FIXED: Ensure this always runs
        for day in range(1, month_days + 1):
            if hasattr(self, 'daily_total_vars') and hasattr(self, 'leave_total_vars'):
                if day in self.daily_total_vars and day in self.leave_total_vars:
                    direct_total = self.daily_total_vars[day].get()
                    leave_total = self.leave_total_vars[day].get()
                    if hasattr(self, 'grand_total_vars') and day in self.grand_total_vars:
                        self.grand_total_vars[day].set(round(direct_total + leave_total, 2))
        
        # Additional: Update project totals in real-time
        self.update_project_total_displays()

    def update_project_total_displays(self):
        """Update project total displays in real-time with progress - IMPROVED"""
        for project_idx, project in enumerate(self.projects):
            total_hours = 0.0
            for day in project['entries']:
                hours_str = project['entries'][day].get()
                total_hours += self.safe_float_convert(hours_str)
            
            # Update the total variable - ensure this always runs
            if 'total_var' in project:
                project['total_var'].set(round(total_hours, 1))
            
            # Update progress percentage
            allocated_hours = self.safe_float_convert(project['hours_var'].get())
            if allocated_hours > 0:
                progress_percent = (total_hours / allocated_hours) * 100
                if 'progress_var' in project:
                    project['progress_var'].set(f"{progress_percent:.0f}%")
                    
                # Color code based on progress
                if 'progress_label' in project:
                    if progress_percent > 100:
                        project['progress_label'].config(foreground=ModernTheme.ACCENT)  # Red for over-allocation
                    elif progress_percent >= 80:
                        project['progress_label'].config(foreground=ModernTheme.WARNING)  # Orange for near limit
                    else:
                        project['progress_label'].config(foreground=ModernTheme.INFO)  # Blue for normal
            else:
                # If no allocated hours, show just the total
                if 'progress_var' in project:
                    project['progress_var'].set("")

    def calculate_totals(self):
        """Calculate all totals for export with safe conversion"""
        year = self.selected_year.get()
        month_name = self.selected_month.get()
        month_index = EthiopianDateConverter.MONTHS_AMHARIC.index(month_name) + 1
        month_days = EthiopianDateConverter.get_ethiopian_month_days(year, month_index)
        
        results = {
            'project_totals': [],
            'daily_totals': {day: 0.0 for day in range(1, month_days + 1)},
            'leave_totals': {day: 0.0 for day in range(1, month_days + 1)},
            'grand_totals': {day: 0.0 for day in range(1, month_days + 1)}
        }
        
        # Calculate project totals with safe conversion
        for project_idx, project in enumerate(self.projects):
            project_total = 0.0
            for day in range(1, month_days + 1):
                hours_var = project['entries'].get(day)
                hours_str = hours_var.get() if hours_var else "0"
                hours = self.safe_float_convert(hours_str)
                project_total += hours
                results['daily_totals'][day] += hours
            
            allocated_hours = self.safe_float_convert(project['hours_var'].get())
            
            total_work_hours = sum(results['daily_totals'].values())
            
            results['project_totals'].append({
                'name': project['name_var'].get(),
                'total_hours': project_total,
                'allocated_hours': allocated_hours,
                'equiv_days': project_total / 8 if project_total > 0 else 0,
                'percent_direct': (project_total / total_work_hours * 100) if total_work_hours > 0 else 0,
                'percent_total': (project_total / (month_days * 8) * 100) if month_days * 8 > 0 else 0
            })
        
        # Calculate leave totals with safe conversion
        for day in range(1, month_days + 1):
            day_leave_total = 0.0
            for leave_key in self.leave_data:
                hours_var = self.leave_data[leave_key]['entries'].get(day)
                hours_str = hours_var.get() if hours_var else "0"
                day_leave_total += self.safe_float_convert(hours_str)
            results['leave_totals'][day] = day_leave_total
        
        # Calculate grand totals
        for day in range(1, month_days + 1):
            results['grand_totals'][day] = results['daily_totals'][day] + results['leave_totals'][day]
        
        return results
    
    def calculate_total_working_hours(self):
        """Calculate total working hours for the selected month based on workdays"""
        year = self.selected_year.get()
        month_name = self.selected_month.get()
        month_index = EthiopianDateConverter.MONTHS_AMHARIC.index(month_name) + 1
        month_days = EthiopianDateConverter.get_ethiopian_month_days(year, month_index)
        
        total_hours = 0.0
        
        for day in range(1, month_days + 1):
            weekday_index = EthiopianDateConverter.get_ethiopian_weekday(year, month_index, day)
            
            # Calculate hours based on weekday (same logic as prefill_default_hours)
            if weekday_index < 4:  # Monday-Thursday: 8 hours
                day_hours = 8.0
            elif weekday_index == 4:  # Friday: 8 hours
                day_hours = 8.0
            elif weekday_index == 5:  # Saturday: 4 hours
                day_hours = 4.0
            else:  # Sunday: 0 hours
                day_hours = 0.0
            
            total_hours += day_hours
        
        return total_hours    

    def preview_timesheet(self):
        """Preview the timesheet in a small window"""
        # REMOVE this check:
        # if not self.name_entered:
        #     messagebox.showerror("Error", "Please enter your full name first before previewing the timesheet.")
        #     return
        
        results = self.calculate_totals()
        
        # Create preview window (smaller size)
        preview = tk.Toplevel(self.root)
        preview.title("Timesheet Preview - የሰዓት ሰሌዳ አጭር/ማጠቃለያ እይታ")
        preview.geometry("800x600")  # Smaller window
        
        # Create text widget for preview
        text_frame = ttk.Frame(preview, style="Modern.TFrame")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = scrolledtext.ScrolledText(text_frame, wrap=tk.WORD, font=("Courier", 9))
        text_widget.pack(fill=tk.BOTH, expand=True)
        
        # Generate preview text
        preview_text = self.generate_preview_text(results)
        text_widget.insert(tk.END, preview_text)
        text_widget.config(state=tk.DISABLED)
        
        # Add close button
        close_btn = ttk.Button(preview, text="Close", command=preview.destroy, style="Primary.TButton")
        close_btn.pack(pady=10)

    def generate_preview_text(self, results):
        """Generate formatted preview text"""
        year = self.selected_year.get()
        month_name = self.selected_month.get()
        month_index = EthiopianDateConverter.MONTHS_AMHARIC.index(month_name) + 1
        month_days = EthiopianDateConverter.get_ethiopian_month_days(year, month_index)
        
        text = "MERQ CONSULTANCY\n"
        text += "=" * 80 + "\n"
        text += "ወርሃዊ የስራ ሰዓት መከታተያ / Monthly Timesheet Tracker \n"
        text += "=" * 80 + "\n\n"
        text += f"ሰራተኛ/አማካሪ ስም / Employee/Consultant Name: {self.employee_name.get()}\n"
        text += f"ወር / Month: {month_name} {year}\n\n"
        
        text += "ፕሮጀክቶች / Projects Summary:\n"
        text += "-" * 80 + "\n"
        for project in results['project_totals']:
            text += f"• {project['name']}:\n"
            text += f"  ጠቅላላ ሰዓቶች / Total Hours: {project['total_hours']:.1f}\n"
            text += f"  የተመደበው ሰዓት / Hours Allocated: {project['allocated_hours']:.1f}\n"
            text += f"  እኩል የተመደበላቸው ቀናት / Equiv. Days Allocated: {project['equiv_days']:.1f}\n"
            text += f"  % OF ቀጥታ: {project['percent_direct']:.1f}%\n"
            text += f"  % OF ጠቅላላ: {project['percent_total']:.1f}%\n\n"
        
        text += "ጠቅላላ ሰዓቶች ማጠቃለያ / Total Hours Summary:\n"
        text += "-" * 80 + "\n"
        
        total_work_hours = sum(results['daily_totals'].values())
        total_leave_hours = sum(results['leave_totals'].values())
        grand_total = total_work_hours + total_leave_hours
        
        text += f"ጠቅላላ የስራ ሰዓቶች / Total Work Hours: {total_work_hours:.1f}\n"
        text += f"ጠቅላላ የፈቃድ ሰዓቶች / Total Leave Hours: {total_leave_hours:.1f}\n"
        text += f"ጠቅላላ ሁሉ / Grand Total: {grand_total:.1f}\n\n"
        
        text += "ይህ የጊዜ ሰሌዳ የሚያገለገልዉ ለ" + month_name + " ወር " + str(year) + " ዓ.ም. ብቻ ነው።\n\n"
        
        text += "Declaration:\n"
        text += "እኔ፣ ከዚህ በላይ ያለው መረጃ እውነት መሆኑን፣ ከእውነታው በኋላ የሚወሰነው እና በእኔ በተከናወነው ትክክለኛ ስራ ላይ የተመሰረተ መሆኑን እገልጻለሁ።\n"
        text += "I, hereby declare that the foregoing information is true, is determined after the fact and is based on actual work performed by me.\n"
        
        return text

    def export_to_excel(self):
        """Export timesheet to Excel format using the template file"""
        # REMOVE this check:
        # if not self.name_entered:
        #     messagebox.showerror("Error", "Please enter your full name first before exporting to Excel.")
        #     return
        
        try:
            results = self.calculate_totals()
            year = self.selected_year.get()
            month_name = self.selected_month.get()
            month_index = EthiopianDateConverter.MONTHS_AMHARIC.index(month_name) + 1
            month_days = EthiopianDateConverter.get_ethiopian_month_days(year, month_index)
            
            # Calculate total allocated hours from all projects
            total_allocated_hours = sum(self.safe_float_convert(project['hours_var'].get()) for project in self.projects)
            
            # Check if template file exists
            template_file = "MERQ_TIMESHEET_ETH-CAL_TEMPLATE.xlsx"
            if not os.path.exists(template_file):
                messagebox.showerror("Error", f"Template file '{template_file}' not found in the current directory")
                return
            
            # Load the template workbook
            workbook = load_workbook(template_file)
            worksheet = workbook.active
            
            # Safe method to update cells (handles merged cells)
            def safe_cell_update(cell_ref, value):
                """Safely update a cell, handling merged cells"""
                try:
                    # Check if cell is part of a merged range
                    cell = worksheet[cell_ref]
                    for merged_range in list(worksheet.merged_cells.ranges):
                        if cell.coordinate in merged_range:
                            # Unmerge the range first
                            worksheet.unmerge_cells(str(merged_range))
                            break
                    worksheet[cell_ref] = value
                    return True
                except Exception as e:
                    print(f"Warning: Could not update cell {cell_ref}: {e}")
                    return False
            
            # Update header content - use safe method
            header_updates = [
                ('AJ19', f"{month_name} {year}"),
                ('C25', f"{month_name} {year}"), 
                ('AJ3', f"{month_name} {year}"),
                ('H5', self.employee_name.get()),
                ('X4', month_name),
                ('X5', year),
                ('AI8', total_allocated_hours)  # Add total allocated hours here
            ]
            
            for cell_ref, value in header_updates:
                safe_cell_update(cell_ref, value)
            
            # Clear existing date data from template (columns D to AG, rows 6-7)
            for day in range(1, 32):  # Clear up to 31 days
                col = get_column_letter(3 + day)
                safe_cell_update(f'{col}6', "")  # Clear date
                safe_cell_update(f'{col}7', "")  # Clear weekday
            
            # Fill in the actual dates and weekdays
            for day in range(1, month_days + 1):
                col = get_column_letter(3 + day)  # Starting from column D
                
                # Fill date (row 6)
                eth_date = EthiopianDateConverter.format_ethiopian_date(year, month_index, day)
                safe_cell_update(f'{col}6', eth_date)
                
                # Fill weekday (row 7)
                weekday_index = EthiopianDateConverter.get_ethiopian_weekday(year, month_index, day)
                weekday_name = EthiopianDateConverter.WEEK_DAYS_AMHARIC[weekday_index]
                safe_cell_update(f'{col}7', weekday_name)
            
            # Clear existing project data
            for row in range(8, 15):  # Rows 8-14 for projects
                for day in range(1, 32):  # Up to 31 days
                    col = get_column_letter(3 + day)
                    safe_cell_update(f'{col}{row}', 0)  # Clear project hours
            
            # Fill project hours
            for project_idx, project in enumerate(self.projects):
                if project_idx >= 7:  # Template only has space for 7 projects
                    break
                    
                row = 8 + project_idx
                
                # Update project name in column B
                safe_cell_update(f'B{row}', project['name_var'].get())
                
                # Fill daily hours
                for day in range(1, month_days + 1):
                    col = get_column_letter(3 + day)
                    hours_var = project['entries'].get(day)
                    hours_str = hours_var.get() if hours_var else "0"
                    hours = self.safe_float_convert(hours_str)
                    
                    if hours > 0:
                        safe_cell_update(f'{col}{row}', hours)
            
            # Clear existing leave data
            for row in range(16, 22):  # Rows 16-21 for leave
                for day in range(1, 32):  # Up to 31 days
                    col = get_column_letter(3 + day)
                    safe_cell_update(f'{col}{row}', 0)  # Clear leave hours
            
            # Fill leave hours
            leave_types = [
                ("የእረፍት ጊዜ / VACATION", "vacation"),
                ("የጤና እረፍት / SICK LEAVE", "sick_leave"),
                ("በዓል / HOLIDAY", "holiday"),
                ("የግል ፈቃድ / PERSONAL LEAVE", "personal_leave"),
                ("የሐዘን እረፍት / BEREAVEMENT", "bereavement"),
                ("ሌሎች / Other", "other")
            ]
            
            for i, (leave_name, leave_key) in enumerate(leave_types):
                if i >= 6:  # Template only has space for 6 leave types
                    break
                    
                row = 16 + i
                
                # Fill daily leave hours
                for day in range(1, month_days + 1):
                    col = get_column_letter(3 + day)
                    hours_var = self.leave_data[leave_key]['entries'].get(day)
                    hours_str = hours_var.get() if hours_var else "0"
                    hours = self.safe_float_convert(hours_str)
                    
                    if hours > 0:
                        safe_cell_update(f'{col}{row}', hours)
            
            # Update signature section
#            current_greg_date = datetime.now().strftime("%d/%m/%Y")
#            safe_cell_update('K29', current_greg_date)  # Employee date
#            safe_cell_update('AJ29', current_greg_date)  # Supervisor date
            
            # Use Ethiopian date for signature section
            eth_year, eth_month, eth_day = EthiopianDateConverter.gregorian_to_ethiopian(datetime.now())  # Pass datetime object

            # Format Ethiopian date as 'day/month/year'
            eth_date_str = f"{eth_day:02d}/{eth_month:02d}/{eth_year}"

            # Update cells with the formatted Ethiopian date
            safe_cell_update('K29', eth_date_str)  # Employee date
            safe_cell_update('AJ29', eth_date_str)  # Supervisor date
                        
            
            # Update the employee name in signature section (row 29)
            safe_cell_update('B29', self.employee_name.get())
            
            # Update employee's supervisor name from session if available
            if self.user_session and self.user_session.supervisor_name:
                safe_cell_update('P29', self.user_session.supervisor_name)
            #safe_cell_update('B29', self.employee_name.get())

            # Update employee's supervisor position from session if available
            if self.user_session and self.user_session.supervisor_position_title:
                safe_cell_update('T29', self.user_session.supervisor_position_title)

            
            # Add timestamp to filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Ask for save location
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"MERQ_TIMESHEET_{self.employee_name.get()}_{month_name}_{year}_{timestamp}.xlsx"
            )
            
            if filename:
                workbook.save(filename)
                messagebox.showinfo("Success", f"Timesheet successfully exported using template:\n{filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Could not export file: {str(e)}\n\nError details: {str(e)}")

    def submit_timesheet(self):
        """Submit timesheet to HR via email"""
        if not self.user_session:
            messagebox.showerror("Error", "Please login first")
            return
        
        if not self.name_entered:
            messagebox.showerror("Error", "Please complete your timesheet first")
            return
        
        # First export to Excel
        temp_file = self.export_to_excel_for_submission()
        if not temp_file:
            return
        
        # Confirm submission
        confirm_msg = f"""
        Are you sure you want to submit your timesheet to HR?
        
        This will send your timesheet to:
        HR Department - haymanot.a@merqconsultancy.org
        
        Employee: {self.user_session.full_name}
        Position: {self.user_session.position}
        Department: {self.user_session.department}
        
        Once submitted, you cannot make changes to this month's timesheet.
        """
        
        if messagebox.askyesno("Confirm Submission", confirm_msg):
            # Send email
            success = self.email_manager.send_timesheet_email(temp_file, self.user_session)
            
            if success:
                messagebox.showinfo("Success", "Timesheet submitted successfully to HR!")
                # Optional: Lock the timesheet for editing
                self.lock_timesheet_editing()
            else:
                messagebox.showerror("Error", "Failed to send email. Please check your email configuration or send manually.")

    def export_to_excel_for_submission(self):
        """Export to Excel for submission (similar to existing export but with user data)"""
        try:
            # Use existing export logic but enhance with user data
            filename = f"MERQ_TIMESHEET_{self.user_session.full_name}_{self.selected_month.get()}_{self.selected_year.get()}_SUBMITTED.xlsx"
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if filename:
                # Use the existing export functionality
                self.export_to_excel()
                return filename
            return None
        except Exception as e:
            messagebox.showerror("Error", f"Could not export file: {str(e)}")
            return None

    def lock_timesheet_editing(self):
        """Lock the timesheet after submission to prevent changes"""
        # This is a placeholder for future implementation
        # You can disable input fields after submission
        pass

    def clear_all(self):
        """Clear all data"""
        if messagebox.askyesno("Confirm", "ሁሉንም ውሂብ ማጽዳት ይፈልጋሉ? / Clear all data?"):
            self.employee_name.set("PLEASE ENTER YOUR FULL NAME")
            self.selected_year.set(self.current_ethiopian_date['year'])
            self.selected_month.set(self.current_ethiopian_date['month_name'])
            
            # Clear projects (keep only default)
            for project in self.projects[1:]:  # Keep first project (MERQ Internal)
                project['frame'].destroy()
            self.projects = self.projects[:1]
            self.daily_hours = {0: {}}
            
            # Clear leave data
            for leave_key in self.leave_data:
                self.leave_data[leave_key] = {'entries': {}, 'total_var': tk.DoubleVar(value=0.0)}
            
            self.update_calendar_display()

    def on_closing(self):
        """Handle application closing with confirmation"""
        if messagebox.askyesno("Confirm Exit", 
                            "Are you sure you want to exit the MERQ Timesheet Application?\n\n"
                            "የመርክ የሰዓት ሰሌዳ መተግበሪያውን ለመውጣት እርግጠኛ ነዎት?"):
            self.root.quit()

def main():
    root = tk.Tk()
    
    # Check if database exists
    if not os.path.exists("merq_timesheet_db.sqlite"):
        messagebox.showerror("Database Error", 
                           "Database file 'merq_timesheet_db.sqlite' not found.\n"
                           "Please ensure the database file is in the same directory as the application.")
        return
    
    app = TimesheetApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()